"""
OUTPUT FORMATTER
================
Contains functions for generating output Excel files.
"""

import os
import time
import pandas as pd
import traceback
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from logging_config import (
    get_logger,
    log_step,
    log_file_operation,
    log_calculation_result,
    log_dataframe_info,
    ProgressTracker,
)
from const import (
    RGA_SHARE_BY_DEAL,
    GROSS_UP_FACTOR_BY_DEAL,
    GROSS_UP_FACTOR_PENSIONER,
    GROSS_UP_FACTOR_DEFERRED,
    get_deal_name_from_rga_tab,
    is_deferred_pensioner_tranche,
    RGA_TO_CLIENT_MAPPING,
)
from client_data_extractor import normalize_date_for_matching

# Initialize logger for this module
logger = get_logger("output_formatter")

# Excel styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


@log_step("Writing inflation-adjusted output")
def write_inflation_adjusted_output(comprehensive_results, output_file):
    """Write detailed inflation-adjusted results to Excel."""
    logger.info(f"Writing inflation-adjusted output to: {output_file}")
    logger.debug(f"Processing {len(comprehensive_results)} tranches")

    # Ensure output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.debug(f"Created output directory: {output_dir}")

    try:
        start_time = time.time()
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            tracker = ProgressTracker(
                total=len(comprehensive_results),
                name="Writing Tranche Sheets",
                log_interval=25,
                logger=logger
            )
            
            for tranche_name, df in comprehensive_results.items():
                df.to_excel(writer, sheet_name=tranche_name, index=False)
                logger.debug(f"Wrote sheet '{tranche_name}': {len(df)} rows")
                tracker.update()

            tracker.complete()

            # Processing log sheet
            log_data = []
            for tranche_name, df in comprehensive_results.items():
                claims_count = df['Total_Actual_Claims'].notna().sum() if 'Total_Actual_Claims' in df.columns else 0
                log_data.append({
                    'Tranche': tranche_name,
                    'Rows': len(df),
                    'Claims_Matched': claims_count
                })

            log_df = pd.DataFrame(log_data)
            log_df.to_excel(writer, sheet_name='Processing_Log', index=False)
            logger.debug(f"Wrote Processing_Log sheet: {len(log_data)} entries")

        elapsed = time.time() - start_time
        logger.info(f"Inflation-adjusted output created in {elapsed:.2f}s")
        log_file_operation(logger, "WRITE", output_file, True, f"{len(comprehensive_results)} sheets")
        return True

    except Exception as e:
        logger.exception(f"Error writing inflation-adjusted output: {str(e)}")
        log_file_operation(logger, "WRITE", output_file, False, str(e))
        return False


@log_step("Creating Val_Date output")
def create_valdate_output(comprehensive_results, rga_discount_curve, client_discount_df,
                          valuation_date, exposure_summary_data, fee_pv_data,
                          fee_vectors_data, output_file):
    """Create the Val_Date Output file with reorganized structure."""
    logger.info(f"Creating Val_Date output: {output_file}")
    logger.debug(f"Valuation date: {valuation_date}")
    logger.debug(f"Tranches to process: {len(comprehensive_results)}")

    # Ensure output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.debug(f"Created output directory: {output_dir}")

    start_time = time.time()

    # Build RGA discount lookup
    rga_discount_lookup = {}
    if rga_discount_curve is not None and not rga_discount_curve.empty:
        for _, row in rga_discount_curve.iterrows():
            rga_discount_lookup[row['Date']] = row.get('Final_Discount_Curve', None)
        logger.debug(f"RGA discount lookup built: {len(rga_discount_lookup)} entries")

    # Build client discount lookup
    client_discount_lookup = {}
    if client_discount_df is not None and not client_discount_df.empty:
        for _, row in client_discount_df.iterrows():
            normalized = normalize_date_for_matching(row['Date'])
            if normalized:
                client_discount_lookup[normalized] = row.get('Client_Discount_Factor', None)
        logger.debug(f"Client discount lookup built: {len(client_discount_lookup)} entries")

    # Group tranches by deal
    deals_data = {}
    tranche_to_full_name = {}

    for tranche_name, df in comprehensive_results.items():
        deal_name = get_deal_name_from_rga_tab(tranche_name)
        if deal_name not in deals_data:
            deals_data[deal_name] = {}

        if tranche_name.startswith(deal_name + "_"):
            tranche_suffix = tranche_name[len(deal_name) + 1:]
        else:
            tranche_suffix = tranche_name

        deals_data[deal_name][tranche_suffix] = df
        tranche_to_full_name[(deal_name, tranche_suffix)] = tranche_name

    logger.info(f"Grouped tranches into {len(deals_data)} deals")
    for deal, tranches in deals_data.items():
        logger.debug(f"  Deal '{deal}': {len(tranches)} tranches")

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

            # Write each deal tab
            deal_tracker = ProgressTracker(
                total=len(deals_data),
                name="Writing Deal Sheets",
                log_interval=25,
                logger=logger
            )
            
            for deal_name, tranches in deals_data.items():
                logger.debug(f"Processing deal: {deal_name}")
                
                rga_share = RGA_SHARE_BY_DEAL.get(deal_name, 1.0)
                gross_up = GROSS_UP_FACTOR_BY_DEAL.get(deal_name, 0.0)
                gross_up_pensioner = GROSS_UP_FACTOR_PENSIONER.get(deal_name, 0.0)
                gross_up_deferred = GROSS_UP_FACTOR_DEFERRED.get(deal_name, 0.0)

                logger.debug(f"  RGA share: {rga_share}, Gross-up: {gross_up}")

                all_dates = set()
                for tranche_suffix, df in tranches.items():
                    df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
                    val_date = pd.to_datetime(valuation_date)
                    filtered_dates = df[df['Date_Datetime'] >= val_date]['Date'].tolist()
                    all_dates.update(filtered_dates)

                all_dates = sorted(list(all_dates))
                if not all_dates:
                    logger.warning(f"No dates found for deal {deal_name} after valuation date")
                    deal_tracker.update()
                    continue

                logger.debug(f"  Date range: {all_dates[0]} to {all_dates[-1]} ({len(all_dates)} dates)")

                result_df = pd.DataFrame({'Date': all_dates})
                result_df['RGA_Discount'] = result_df['Date'].apply(lambda x: rga_discount_lookup.get(x, None))
                result_df['Client_Discount'] = result_df['Date'].apply(lambda x: client_discount_lookup.get(x, None))

                current_col = 4
                tranche_columns = {}

                for tranche_suffix in sorted(tranches.keys()):
                    df = tranches[tranche_suffix]
                    full_tranche_name = tranche_to_full_name.get((deal_name, tranche_suffix), f"{deal_name}_{tranche_suffix}")
                    is_deferred_pensioner = is_deferred_pensioner_tranche(full_tranche_name)

                    df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
                    val_date = pd.to_datetime(valuation_date)
                    df_filtered = df[df['Date_Datetime'] >= val_date].copy()

                    fixed_lookup = dict(zip(df_filtered['Date'], df_filtered['Fixed_w_Real']))
                    float_lookup = dict(zip(df_filtered['Date'], df_filtered['Adjusted_Float_Vector']))

                    fixed_col_letter = get_column_letter(current_col)
                    result_df[f'{tranche_suffix}_Fixed_w_Real'] = result_df['Date'].apply(
                        lambda x, fl=fixed_lookup: fl.get(x, None)
                    )

                    float_col_letter = get_column_letter(current_col + 1)
                    result_df[f'{tranche_suffix}_Adj_Float_Vector'] = result_df['Date'].apply(
                        lambda x, fl=float_lookup: fl.get(x, None)
                    )

                    if is_deferred_pensioner:
                        deferred_lookup = dict(zip(df_filtered['Date'], df_filtered.get('Deferred_w_Real', pd.Series())))
                        pensioner_lookup = dict(zip(df_filtered['Date'], df_filtered.get('Pensioner_w_Real', pd.Series())))

                        deferred_col_letter = get_column_letter(current_col + 2)
                        result_df[f'{tranche_suffix}_Deferred_w_Real'] = result_df['Date'].apply(
                            lambda x, fl=deferred_lookup: fl.get(x, None)
                        )

                        pensioner_col_letter = get_column_letter(current_col + 3)
                        result_df[f'{tranche_suffix}_Pensioner_w_Real'] = result_df['Date'].apply(
                            lambda x, fl=pensioner_lookup: fl.get(x, None)
                        )

                        fee_col_letter = get_column_letter(current_col + 4)
                        result_df[f'{tranche_suffix}_Fee'] = None

                        tranche_columns[tranche_suffix] = {
                            'fixed_col': fixed_col_letter,
                            'float_col': float_col_letter,
                            'deferred_col': deferred_col_letter,
                            'pensioner_col': pensioner_col_letter,
                            'fee_col': fee_col_letter,
                            'is_deferred_pensioner': True
                        }
                        current_col += 5
                    else:
                        fee_col_letter = get_column_letter(current_col + 2)
                        result_df[f'{tranche_suffix}_Fee'] = None

                        tranche_columns[tranche_suffix] = {
                            'fixed_col': fixed_col_letter,
                            'float_col': float_col_letter,
                            'fee_col': fee_col_letter,
                            'is_deferred_pensioner': False
                        }
                        current_col += 3

                sheet_name = deal_name[:31] if len(deal_name) > 31 else deal_name
                result_df.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]

                # Add Fee formulas
                for tranche_suffix, cols in tranche_columns.items():
                    fee_col = cols['fee_col']

                    if cols['is_deferred_pensioner']:
                        deferred_col = cols['deferred_col']
                        pensioner_col = cols['pensioner_col']
                        for row in range(2, len(result_df) + 2):
                            ws[f'{fee_col}{row}'] = (
                                f'={deferred_col}{row}*{gross_up_deferred}*{rga_share}+'
                                f'{pensioner_col}{row}*{gross_up_pensioner}*{rga_share}'
                            )
                    else:
                        fixed_col = cols['fixed_col']
                        for row in range(2, len(result_df) + 2):
                            ws[f'{fee_col}{row}'] = f'={fixed_col}{row}*{rga_share}*{gross_up}'

                # Format header
                for cell in ws[1]:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal='center')

                ws.freeze_panes = 'A2'

                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 20)

                logger.info(f"Wrote deal sheet '{sheet_name}': {len(result_df)} rows, {len(tranche_columns)} tranches")
                deal_tracker.update()

            deal_tracker.complete()

            # Client Exposure Summary
            if exposure_summary_data or fee_pv_data:
                logger.debug("Writing Client Exposure Summary sheet")
                summary_list = list(exposure_summary_data) if exposure_summary_data else []
                
                # Add Fee tranches for deals with fee vectors
                for deal_name, fee_pv_value in fee_pv_data.items():
                    summary_list.append({
                        'Deal': deal_name.replace('_', ' ').title(),  # Format: "Advance T1a"
                        'Tranche': 'Fee',
                        'Fee_PV': fee_pv_value,
                        'A_PV': None,
                        'E_PV': None,
                        'Exposure': None,
                        'Experience_Factor': None
                    })
                
                summary_df = pd.DataFrame(summary_list)
                summary_df = summary_df[['Deal', 'Tranche', 'Fee_PV', 'A_PV', 'E_PV', 'Experience_Factor', 'Exposure']]
                summary_df.to_excel(writer, sheet_name='Client Exposure Summary', index=False)

                ws_summary = writer.sheets['Client Exposure Summary']
                for cell in ws_summary[1]:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal='center')
                ws_summary.freeze_panes = 'A2'
                logger.info(f"Wrote Client Exposure Summary: {len(summary_list)} entries")

            # RGA Summary
            logger.debug("Building RGA Summary sheet")
            rga_summary_data = _build_rga_summary(deals_data, tranche_to_full_name, valuation_date,
                                                   rga_discount_lookup, client_discount_lookup,
                                                   fee_vectors_data, fee_pv_data)

            if rga_summary_data:
                rga_summary_df = pd.DataFrame(rga_summary_data)
                
                # Standard columns (Experience Factor after Tranche)
                output_cols = ['Deal', 'Tranche', 'Experience_Factor', 'RGA_Share', 'RGA_PV_Fixed_Vector', 'RGA_PV_Floating_Vector',
                             'RGA_PV_Fee', 'Client_PV_Fixed_Vector', 'Client_PV_Floating_Vector', 'Client_PV_Fee']
                
                # Only keep columns that exist in the dataframe
                output_cols = [col for col in output_cols if col in rga_summary_df.columns]
                rga_summary_df = rga_summary_df[output_cols]
                
                rga_summary_df.to_excel(writer, sheet_name='RGA Summary', index=False, header=False, startrow=2)

                ws_rga = writer.sheets['RGA Summary']

                ws_rga['A1'] = 'Deal'
                ws_rga['B1'] = 'Tranche'
                ws_rga['C1'] = 'Experience Factor'
                ws_rga['D1'] = 'RGA Share'
                ws_rga['E1'] = 'RGA Discount'
                ws_rga['H1'] = 'Client Discount'

                ws_rga.merge_cells('A1:A2')
                ws_rga.merge_cells('B1:B2')
                ws_rga.merge_cells('C1:C2')
                ws_rga.merge_cells('D1:D2')
                ws_rga.merge_cells('E1:G1')
                ws_rga.merge_cells('H1:J1')

                ws_rga['E2'] = 'PV of Fixed Vector'
                ws_rga['F2'] = 'PV of Floating Vector'
                ws_rga['G2'] = 'PV Fee'
                ws_rga['H2'] = 'PV of Fixed Vector'
                ws_rga['I2'] = 'PV of Floating Vector'
                ws_rga['J2'] = 'PV Fee'

                for row in [1, 2]:
                    for cell in ws_rga[row]:
                        cell.fill = HEADER_FILL
                        cell.font = HEADER_FONT
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                ws_rga.freeze_panes = 'A3'
                logger.info(f"Wrote RGA Summary: {len(rga_summary_data)} entries")

        elapsed = time.time() - start_time
        logger.info(f"Val_Date output created in {elapsed:.2f}s")
        log_file_operation(logger, "WRITE", output_file, True, f"{len(deals_data)} deal sheets")
        return True

    except Exception as e:
        logger.exception(f"Error creating Val_Date output: {str(e)}")
        log_file_operation(logger, "WRITE", output_file, False, str(e))
        return False


def _build_rga_summary(deals_data, tranche_to_full_name, valuation_date, rga_discount_lookup,
                       client_discount_lookup, fee_vectors_data, fee_pv_data):
    """Build RGA summary data with PV calculations."""
    from client_data_extractor import normalize_date_for_matching
  
    logger.debug("Building RGA summary data")
    rga_summary_data = []

    for deal_name, tranches in deals_data.items():
        rga_share = RGA_SHARE_BY_DEAL.get(deal_name, 1.0)
        gross_up = GROSS_UP_FACTOR_BY_DEAL.get(deal_name, 0.0)
        gross_up_pensioner = GROSS_UP_FACTOR_PENSIONER.get(deal_name, 0.0)
        gross_up_deferred = GROSS_UP_FACTOR_DEFERRED.get(deal_name, 0.0)

        # Process regular tranches
        for tranche_suffix in sorted(tranches.keys()):
            df = tranches[tranche_suffix]
            full_tranche_name = tranche_to_full_name.get((deal_name, tranche_suffix), f"{deal_name}_{tranche_suffix}")
            is_dp = is_deferred_pensioner_tranche(full_tranche_name)

            df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
            val_date = pd.to_datetime(valuation_date)
            df_filtered = df[df['Date_Datetime'] >= val_date].copy()

            if df_filtered.empty:
                logger.debug(f"No data after valuation date for {full_tranche_name}")
                continue

            fixed_w_real = pd.to_numeric(df_filtered['Fixed_w_Real'], errors='coerce').fillna(0)
            adj_float_vector = pd.to_numeric(df_filtered['Adjusted_Float_Vector'], errors='coerce').fillna(0)

            if is_dp:
                deferred_w_real = pd.to_numeric(df_filtered.get('Deferred_w_Real', pd.Series()), errors='coerce').fillna(0)
                pensioner_w_real = pd.to_numeric(df_filtered.get('Pensioner_w_Real', pd.Series()), errors='coerce').fillna(0)
                fee = deferred_w_real * gross_up_deferred * rga_share + pensioner_w_real * gross_up_pensioner * rga_share
            else:
                fee = fixed_w_real * rga_share * gross_up

            rga_discounts = df_filtered['Date'].apply(
                lambda x: rga_discount_lookup.get(x, 0) if rga_discount_lookup.get(x) is not None else 0
            ).astype(float)

            client_discounts = df_filtered['Date'].apply(
                lambda x: client_discount_lookup.get(x, 0) if client_discount_lookup.get(x) is not None else 0
            ).astype(float)

            rga_pv_fixed = (fixed_w_real * rga_discounts).sum()
            rga_pv_float = (adj_float_vector * rga_discounts).sum()
            rga_pv_fee = (fee * rga_discounts).sum()
            client_pv_fixed = (fixed_w_real * client_discounts).sum()
            client_pv_float = (adj_float_vector * client_discounts).sum()
            client_pv_fee = (fee * client_discounts).sum()

            experience_factor = df['Experience_Factor'].iloc[0] if 'Experience_Factor' in df.columns and len(df) > 0 else None

            rga_summary_data.append({
                'Deal': deal_name,
                'Tranche': tranche_suffix,
                'Experience_Factor': experience_factor,
                'RGA_Share': rga_share,
                'RGA_PV_Fixed_Vector': rga_pv_fixed,
                'RGA_PV_Floating_Vector': rga_pv_float,
                'RGA_PV_Fee': rga_pv_fee,
                'Client_PV_Fixed_Vector': client_pv_fixed,
                'Client_PV_Floating_Vector': client_pv_float,
                'Client_PV_Fee': client_pv_fee
            })
            
            logger.debug(f"RGA Summary for {deal_name}/{tranche_suffix}: "
                        f"RGA_PV_Fixed={rga_pv_fixed:.2f}, RGA_PV_Float={rga_pv_float:.2f}")

    # Add separate Fee_VectorAm and Fee_VectorIn tranches for deals with fee vectors
    for deal_name, fee_data_by_type in fee_vectors_data.items():
        rga_share = RGA_SHARE_BY_DEAL.get(deal_name, 1.0)
        
        for fee_type, fee_df in fee_data_by_type.items():
            if fee_df is None or fee_df.empty:
                continue
            
            # Normalize dates for matching
            fee_df_copy = fee_df.copy()
            fee_df_copy['Date'] = fee_df_copy['Date'].apply(normalize_date_for_matching)
            
            # Filter to dates >= valuation date
            fee_df_copy['Date_Datetime'] = pd.to_datetime(fee_df_copy['Date'], errors='coerce')
            val_date = pd.to_datetime(valuation_date)
            fee_df_filtered = fee_df_copy[fee_df_copy['Date_Datetime'] >= val_date].copy()
            
            if fee_df_filtered.empty:
                continue
            
            # Get fee vector values
            fee_vectors = pd.to_numeric(fee_df_filtered['Fee_Vector'], errors='coerce').fillna(0)
            
            # Get discount factors
            rga_discounts = fee_df_filtered['Date'].apply(
                lambda x: rga_discount_lookup.get(x, 0) if rga_discount_lookup.get(x) is not None else 0
            ).astype(float)
            
            client_discounts = fee_df_filtered['Date'].apply(
                lambda x: client_discount_lookup.get(x, 0) if client_discount_lookup.get(x) is not None else 0
            ).astype(float)
            
            # Calculate PVs
            rga_pv = (fee_vectors * rga_discounts).sum()
            client_pv = (fee_vectors * client_discounts).sum()
            
            # Add as separate tranche (fee PVs go under PV Fee, not PV of Fixed Vector)
            rga_summary_data.append({
                'Deal': deal_name,
                'Tranche': f'Fee_Vector{fee_type}',  # Fee_VectorAm or Fee_VectorIn
                'Experience_Factor': None,
                'RGA_Share': rga_share,
                'RGA_PV_Fixed_Vector': None,
                'RGA_PV_Floating_Vector': None,
                'RGA_PV_Fee': rga_pv,
                'Client_PV_Fixed_Vector': None,
                'Client_PV_Floating_Vector': None,
                'Client_PV_Fee': client_pv
            })
            
            logger.debug(f"Fee vector summary for {deal_name}/Fee_Vector{fee_type}: RGA_PV={rga_pv:.2f}")

    logger.debug(f"Built RGA summary with {len(rga_summary_data)} entries")
    return rga_summary_data


@log_step("Writing sensitivity output")
def write_sensitivity_output(sensitivity_rows, output_file):
    """Write sensitivity results to Excel (one sheet per deal with 6 scenarios)."""
    logger.info(f"Writing sensitivity output to: {output_file}")
    
    if not sensitivity_rows:
        logger.warning("No sensitivity rows to write - skipping file creation")
        return False

    log_calculation_result(logger, "Sensitivity Rows to Write", len(sensitivity_rows))

    # Ensure output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.debug(f"Created output directory: {output_dir}")

    try:
        start_time = time.time()
        df = pd.DataFrame(sensitivity_rows)

        # Apply column adjustments
        df = df.rename(columns={
            "Deal": "Deal name",
            "Sensitivity_Offset": "Date"
        })
        df.insert(1, "Currency", "GBP")

        # Count unique deals and scenarios
        unique_deals = df["Deal name"].nunique() if "Deal name" in df.columns else 0
        unique_scenarios = df["Scenario"].nunique() if "Scenario" in df.columns else 0
        logger.debug(f"Writing {unique_deals} deals, {unique_scenarios} scenarios")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            deal_tracker = ProgressTracker(
                total=unique_deals,
                name="Writing Sensitivity Deal Sheets",
                log_interval=25,
                logger=logger
            )
            
            for deal_name, deal_df in df.groupby("Deal name"):
                sheet = deal_name[:31]
                deal_df.to_excel(writer, sheet_name=sheet, index=False)
                logger.info(f"Wrote sensitivity sheet '{deal_name}': {len(deal_df)} rows")
                deal_tracker.update()
            
            deal_tracker.complete()
        
        elapsed = time.time() - start_time
        logger.info(f"Sensitivity output created in {elapsed:.2f}s")
        log_file_operation(logger, "WRITE", output_file, True, f"{unique_deals} deal sheets")
        return True
        
    except Exception as e:
        logger.exception(f"Error writing sensitivity output: {str(e)}")
        log_file_operation(logger, "WRITE", output_file, False, str(e))
        return False
