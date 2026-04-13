"""
CALCULATION MODULE
==================
Contains all calculation logic for inflation adjustments, discount curves, and experience factors.
"""

import pandas as pd
import numpy as np
import os
import time
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

from logging_config import (
    get_logger,
    log_calculation_result,
    log_dataframe_info,
    log_dict_summary,
    log_step,
    ProgressTracker,
)
from const import (
    DISCOUNT_CURVE_CONFIG,
    SONIA_SHEET_NAME,
    OUTPUT_COLUMNS_ORDER,
    is_deferred_pensioner_tranche,
    get_deal_name_from_rga_tab,
    PROPHET_DEAL_NAME_MAPPING,
    PROPHET_SHEETS,
    SENSITIVITY_SCENARIOS,
    SENSITIVITY_MAX_MONTHS,
    get_rga_share,
    RGA_SHARE_BY_DEAL,
    GROSS_UP_FACTOR_PENSIONER,
    GROSS_UP_FACTOR_BY_DEAL,
    GROSS_UP_FACTOR_DEFERRED,
)
from client_data_extractor import normalize_date_for_matching

# Initialize logger for this module
logger = get_logger("calculation")


@log_step("Loading SONIA rates")
def load_sonia_rates(sonia_file_path, valuation_date):
    """Load SONIA rates from the Monthly Swap Rates Summary file."""
    logger.info(f"Loading SONIA rates from: {sonia_file_path}")
    logger.debug(f"Valuation date: {valuation_date}")

    if not os.path.exists(sonia_file_path):
        logger.error(f"SONIA file not accessible: {sonia_file_path}")
        return {}

    try:
        wb = load_workbook(sonia_file_path, data_only=True)
        logger.debug(f"Workbook loaded with {len(wb.sheetnames)} sheets")

        if SONIA_SHEET_NAME not in wb.sheetnames:
            logger.error(f"Sheet '{SONIA_SHEET_NAME}' not found in workbook")
            logger.debug(f"Available sheets: {wb.sheetnames}")
            wb.close()
            return {}

        ws = wb[SONIA_SHEET_NAME]
        val_date = pd.to_datetime(valuation_date)
        logger.debug(f"Looking for column matching date: {val_date.strftime('%Y-%m')}")

        val_col = None
        header_row = 1

        for col in range(28, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value
            if cell_val is not None:
                try:
                    cell_date = pd.to_datetime(cell_val)
                    if cell_date.year == val_date.year and cell_date.month == val_date.month:
                        val_col = col
                        logger.debug(f"Found valuation date column at column {col}")
                        break
                except:
                    pass

        if val_col is None:
            logger.error("Valuation date column not found in SONIA sheet")
            wb.close()
            return {}

        sonia_rates = {}
        rows_processed = 0
        for row in range(2, min(ws.max_row + 1, DISCOUNT_CURVE_CONFIG["MAX_MONTHS"] + 10)):
            month_val = ws.cell(row=row, column=28).value
            if month_val is None:
                continue

            try:
                month_offset = int(month_val)
            except:
                continue

            rate_val = ws.cell(row=row, column=val_col).value
            if rate_val is not None:
                try:
                    sonia_rates[month_offset] = float(rate_val)
                    rows_processed += 1
                except:
                    pass

        wb.close()
        
        logger.info(f"Successfully loaded {len(sonia_rates)} SONIA rates")
        if sonia_rates:
            logger.debug(f"SONIA rate range: months {min(sonia_rates.keys())} to {max(sonia_rates.keys())}")
            # Log sample rates (first 3 and last 3)
            sorted_months = sorted(sonia_rates.keys())
            sample_rates = {m: sonia_rates[m] for m in sorted_months[:3] + sorted_months[-3:]}
            logger.debug(f"Sample SONIA rates: {sample_rates}")
        
        return sonia_rates

    except Exception as e:
        logger.exception(f"Error loading SONIA rates: {str(e)}")
        return {}


@log_step("Building RGA discount curve")
def build_rga_discount_curve(valuation_date, sonia_rates, spread=0.0, sensitivity_shock=0.0):
    """Build RGA discount curve from SONIA rates with month-end dates."""
    base_adjustment = DISCOUNT_CURVE_CONFIG["BASE_ADJUSTMENT"]
    max_months = DISCOUNT_CURVE_CONFIG["MAX_MONTHS"]
    val_date = pd.to_datetime(valuation_date)

    logger.info(f"Building discount curve for valuation date: {valuation_date}")
    logger.debug(f"Parameters: base_adjustment={base_adjustment}, spread={spread}, "
                f"sensitivity_shock={sensitivity_shock}, max_months={max_months}")

    discount_curve_data = []

    for month in range(0, max_months + 1):
        date = val_date + relativedelta(months=month)
        
        # Normalize to month-end date (same as tranche dates)
        month_end_date = date.to_period('M').to_timestamp('M')

        if month == 0:
            discount_curve_data.append({
                'Month': month,
                'Date': month_end_date.strftime('%Y-%m-%d'),
                'Final_Discount_Curve': 1.0
            })
            continue

        sonia_rate = sonia_rates.get(month, None)
        if sonia_rate is None:
            available_months = [m for m in sonia_rates.keys() if m <= month]
            if available_months:
                sonia_rate = sonia_rates[max(available_months)]
            else:
                sonia_rate = 0.04

        sonia_annual_adj = sonia_rate + base_adjustment + spread + sensitivity_shock
        sonia_monthly_adj = (1 + sonia_annual_adj) ** (1/12) - 1
        final_discount = 1 / ((1 + sonia_monthly_adj) ** month)

        discount_curve_data.append({
            'Month': month,
            'Date': month_end_date.strftime('%Y-%m-%d'),
            'Final_Discount_Curve': final_discount
        })

    result_df = pd.DataFrame(discount_curve_data)
    
    logger.info(f"Discount curve built with {len(result_df)} points")
    # Log key discount factors
    if len(result_df) > 0:
        sample_months = [0, 12, 60, 120, min(240, len(result_df)-1)]
        for m in sample_months:
            if m < len(result_df):
                row = result_df.iloc[m]
                logger.debug(f"  Month {int(row['Month'])}: Date={row['Date']}, DF={row['Final_Discount_Curve']:.6f}")
    
    return result_df


def smart_date_matching(fixed_df, client_df):
    """Smart date matching that handles business day vs calendar day differences."""
    logger.debug(f"Attempting date matching: fixed_df has {len(fixed_df)} rows, client_df has {len(client_df)} rows")
    
    merged_df = pd.merge(
        fixed_df,
        client_df[['Date_Normalized', 'Historical_Infl_Factors', 'Projected_Infl_Factors', 'Discount_Factors_A_E']],
        on='Date_Normalized',
        how='inner'
    )

    if len(merged_df) > 0:
        logger.debug(f"Direct merge successful: {len(merged_df)} rows matched")
        return merged_df

    logger.debug("Direct merge failed, attempting fuzzy date matching...")
    
    fixed_df_dates = fixed_df.copy()
    client_df_dates = client_df.copy()
    fixed_df_dates['Date_Datetime'] = pd.to_datetime(fixed_df_dates['Date_Normalized'], errors='coerce')
    client_df_dates['Date_Datetime'] = pd.to_datetime(client_df_dates['Date_Normalized'], errors='coerce')

    matched_rows = []

    for _, fixed_row in fixed_df_dates.iterrows():
        fixed_date = fixed_row['Date_Datetime']
        if pd.isna(fixed_date):
            continue

        best_match = None
        min_diff = float('inf')

        for _, client_row in client_df_dates.iterrows():
            client_date = client_row['Date_Datetime']
            if pd.isna(client_date):
                continue

            if (fixed_date.year == client_date.year and fixed_date.month == client_date.month):
                day_diff = abs((fixed_date - client_date).days)
                if day_diff <= 3 and day_diff < min_diff:
                    min_diff = day_diff
                    best_match = client_row

        if best_match is not None:
            matched_row = fixed_row.copy()
            matched_row['Historical_Infl_Factors'] = best_match['Historical_Infl_Factors']
            matched_row['Projected_Infl_Factors'] = best_match['Projected_Infl_Factors']
            matched_row['Discount_Factors_A_E'] = best_match['Discount_Factors_A_E']
            matched_rows.append(matched_row)

    result = pd.DataFrame(matched_rows) if matched_rows else pd.DataFrame()
    logger.debug(f"Fuzzy matching result: {len(result)} rows matched")
    return result


def _get_month_end_timestamp(date_str):
    """Convert a date-like string to a month-end Timestamp."""
    ts = pd.to_datetime(date_str, errors='coerce')
    if pd.isna(ts):
        return pd.NaT
    return ts.to_period('M').to_timestamp('M')


def _months_difference(start_date, end_date):
    """Whole-month difference between two Timestamps."""
    if pd.isna(start_date) or pd.isna(end_date):
        return 0
    return (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)


def _get_value_for_month_offset(df, base_month_end, months_offset, column, fallback_to_first=False):
    """Get a single column value for the row at base_month_end + months_offset months."""
    if df.empty:
        return None

    df_local = df.copy()
    df_local['Date_Datetime'] = pd.to_datetime(df_local['Date'], errors='coerce')
    df_local = df_local.sort_values('Date_Datetime')

    target_month = (base_month_end.to_period('M') + months_offset).to_timestamp('M')
    target_row = df_local.loc[df_local['Date_Datetime'] == target_month]

    if not target_row.empty:
        return pd.to_numeric(target_row.iloc[0][column], errors='coerce')

    if fallback_to_first:
        non_null = df_local.dropna(subset=[column])
        if not non_null.empty:
            return pd.to_numeric(non_null.iloc[0][column], errors='coerce')

    return None


def _safe_ratio(numerator, denominator):
    """Safe division that returns None if denominator is zero or any side is NaN."""
    num = pd.to_numeric(pd.Series([numerator]), errors='coerce').iloc[0]
    den = pd.to_numeric(pd.Series([denominator]), errors='coerce').iloc[0]
    if pd.isna(num) or pd.isna(den) or den == 0:
        return None
    return float(num) / float(den)


def compute_experience_and_adjustment_factors(comprehensive_final, valuation_date):
    """Compute experience, interpolation, adjustment, and credibility factors."""
    logger.debug(f"Computing experience factors for valuation date: {valuation_date}")
    
    if comprehensive_final.empty:
        logger.warning("Empty dataframe provided - returning default factors")
        return {
            'Experience_Factor': 0.0,
            'Interpolation_Vector': 0,
            'Adjustment_Factor': 1.0,
            'Credibility_Factor': 0.0,
            'X_Actual': None,
            'X_Fixed': None,
            'X_Increased': None,
            'X_Decreased': None,
        }

    df = comprehensive_final.copy()
    df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('Date_Datetime')

    valuation_month_end = _get_month_end_timestamp(valuation_date)
    inception_date = df['Date_Datetime'].min()

    months_since_inception = _months_difference(inception_date, valuation_month_end) + 1
    raw_months = max(months_since_inception - 6, 0)
    capped_months = min(raw_months, 60)
    credibility_factor = (capped_months / 60.0) ** 2 if capped_months > 0 else 0.0

    logger.debug(f"Months since inception: {months_since_inception}, capped: {capped_months}, "
                f"credibility_factor: {credibility_factor:.4f}")

    c66_actual = _get_value_for_month_offset(df, valuation_month_end, -66, 'Total_Actual_Claims', fallback_to_first=True)
    c6_actual = _get_value_for_month_offset(df, valuation_month_end, -6, 'Total_Actual_Claims', fallback_to_first=False)
    c66_fixed = _get_value_for_month_offset(df, valuation_month_end, -66, 'Fixed_w_Real', fallback_to_first=True)
    c6_fixed = _get_value_for_month_offset(df, valuation_month_end, -6, 'Fixed_w_Real', fallback_to_first=False)
    c66_increased = _get_value_for_month_offset(df, valuation_month_end, -66, 'Increased_w_Real', fallback_to_first=True)
    c6_increased = _get_value_for_month_offset(df, valuation_month_end, -6, 'Increased_w_Real', fallback_to_first=False)
    c66_decreased = _get_value_for_month_offset(df, valuation_month_end, -66, 'Decreased_w_Real', fallback_to_first=True)
    c6_decreased = _get_value_for_month_offset(df, valuation_month_end, -6, 'Decreased_w_Real', fallback_to_first=False)

    numerator_actual = c66_actual - c6_actual if c66_actual is not None and c6_actual is not None else None
    x_actual = _safe_ratio(numerator_actual, c66_actual)
    numerator_fixed = c66_fixed - c6_fixed if c66_fixed is not None and c6_fixed is not None else None
    x_fixed = _safe_ratio(numerator_fixed, c66_fixed)
    numerator_increased = c66_increased - c6_increased if c66_increased is not None and c6_increased is not None else None
    x_increased = _safe_ratio(numerator_increased, c66_increased)
    numerator_decreased = c66_decreased - c6_decreased if c66_decreased is not None and c6_decreased is not None else None
    x_decreased = _safe_ratio(numerator_decreased, c66_decreased)

    experience_raw = 0.0
    if all(v is not None for v in [x_actual, x_fixed, x_increased, x_decreased]):
        if abs(x_actual - x_fixed) < 1e-6:
            experience_raw = 0.0
        elif x_fixed < x_actual:
            denom = (x_fixed - x_decreased)
            experience_raw = 0.0 if abs(denom) < 1e-6 else (x_fixed - x_actual) / denom
        else:
            denom = (x_increased - x_fixed)
            experience_raw = 0.0 if abs(denom) < 1e-6 else (x_actual - x_fixed) / denom

    experience_factor = experience_raw * credibility_factor

    if x_fixed is None or x_actual is None:
        interpolation_vector = 0
    else:
        interpolation_vector = 0 if x_fixed < x_actual else 1

    offsets = [-8, -7, -6]
    sum_claims = 0.0
    sum_fixed_real = 0.0
    sum_increased_real = 0.0
    sum_decreased_real = 0.0

    for offset in offsets:
        month_end = (valuation_month_end.to_period('M') + offset).to_timestamp('M')
        row = df.loc[df['Date_Datetime'] == month_end]
        if row.empty:
            continue
        r = row.iloc[0]
        sum_claims += pd.to_numeric(r.get('Total_Actual_Claims', 0.0), errors='coerce') or 0.0
        sum_fixed_real += pd.to_numeric(r.get('Fixed_w_Real', 0.0), errors='coerce') or 0.0
        sum_increased_real += pd.to_numeric(r.get('Increased_w_Real', 0.0), errors='coerce') or 0.0
        sum_decreased_real += pd.to_numeric(r.get('Decreased_w_Real', 0.0), errors='coerce') or 0.0

    if sum_claims == 0 and sum_fixed_real == 0 and sum_increased_real == 0 and sum_decreased_real == 0:
        adjustment_factor = 1.0
    else:
        base_denominator = (1.0 - experience_factor) * sum_fixed_real
        if interpolation_vector == 0:
            base_denominator += experience_factor * sum_decreased_real
        else:
            base_denominator += experience_factor * sum_increased_real

        adjustment_factor = sum_claims / base_denominator if base_denominator != 0 else 1.0

    result = {
        'Experience_Factor': experience_factor,
        'Interpolation_Vector': interpolation_vector,
        'Adjustment_Factor': adjustment_factor,
        'Credibility_Factor': credibility_factor,
        'X_Actual': x_actual,
        'X_Fixed': x_fixed,
        'X_Increased': x_increased,
        'X_Decreased': x_decreased,
    }
    
    logger.debug(f"Computed factors: experience={experience_factor:.6f}, "
                f"interpolation={interpolation_vector}, adjustment={adjustment_factor:.6f}")
    
    return result


def process_single_tranche(rga_tab_name, fixed_df, client_vectors_data, client_claims_data, valuation_date):
    """Process a single tranche and return comprehensive results."""
    from const import get_rga_share

    logger.debug(f"Processing tranche: {rga_tab_name}")
    start_time = time.time()

    rga_share = get_rga_share(rga_tab_name)
    logger.debug(f"RGA share for {rga_tab_name}: {rga_share}")

    if rga_tab_name not in client_vectors_data:
        logger.warning(f"No client vectors data found for tranche: {rga_tab_name}")
        return None

    client_df = client_vectors_data[rga_tab_name]
    logger.debug(f"Client vectors data: {len(client_df)} rows")

    fixed_df_clean = fixed_df.copy()
    fixed_df_clean['Date_Normalized'] = fixed_df_clean['Date'].apply(normalize_date_for_matching)

    client_df_clean = client_df.copy()
    client_df_clean['Date_Normalized'] = client_df_clean['Date'].apply(normalize_date_for_matching)

    merged_df = smart_date_matching(fixed_df_clean, client_df_clean)

    if len(merged_df) == 0:
        logger.warning(f"No date matches found for tranche: {rga_tab_name}")
        return None

    logger.debug(f"Merged data: {len(merged_df)} rows")
    comprehensive_df = merged_df.copy()

    if rga_tab_name in client_claims_data:
        client_claims_df = client_claims_data[rga_tab_name]
        claims_df_clean = client_claims_df.copy()
        claims_df_clean['Date_Normalized'] = claims_df_clean['Date'].apply(normalize_date_for_matching)
        comprehensive_df = pd.merge(
            comprehensive_df,
            claims_df_clean[['Date_Normalized', 'Total_Actual_Claims']],
            on='Date_Normalized',
            how='left'
        )
        logger.debug(f"Claims data merged: {comprehensive_df['Total_Actual_Claims'].notna().sum()} non-null values")
    else:
        comprehensive_df['Total_Actual_Claims'] = None
        logger.debug("No claims data available for this tranche")

    # Standard calculations
    comprehensive_df['Fixed_w_Original'] = (
        pd.to_numeric(comprehensive_df['Fixed_Vectors'], errors='coerce') *
        pd.to_numeric(comprehensive_df['Historical_Infl_Factors'], errors='coerce')
    )
    comprehensive_df['Increased_w_Original'] = (
        pd.to_numeric(comprehensive_df['Increase_Vectors'], errors='coerce') *
        pd.to_numeric(comprehensive_df['Historical_Infl_Factors'], errors='coerce')
    )
    comprehensive_df['Decreased_w_Original'] = (
        pd.to_numeric(comprehensive_df['Decrease_Vectors'], errors='coerce') *
        pd.to_numeric(comprehensive_df['Historical_Infl_Factors'], errors='coerce')
    )

    comprehensive_df['Fixed_w_Real'] = (
        pd.to_numeric(comprehensive_df['Fixed_Vectors'], errors='coerce') *
        pd.to_numeric(comprehensive_df['Projected_Infl_Factors'], errors='coerce')
    )
    comprehensive_df['Increased_w_Real'] = (
        pd.to_numeric(comprehensive_df['Increase_Vectors'], errors='coerce') *
        pd.to_numeric(comprehensive_df['Projected_Infl_Factors'], errors='coerce')
    )
    comprehensive_df['Decreased_w_Real'] = (
        pd.to_numeric(comprehensive_df['Decrease_Vectors'], errors='coerce') *
        pd.to_numeric(comprehensive_df['Projected_Infl_Factors'], errors='coerce')
    )

    # Deferred/Pensioner calculations for special tranches
    is_dp_tranche = is_deferred_pensioner_tranche(rga_tab_name)
    if is_dp_tranche:
        logger.debug(f"Tranche {rga_tab_name} is deferred/pensioner type")
        if 'Fixed_deferred' in comprehensive_df.columns:
            comprehensive_df['Deferred_w_Real'] = (
                pd.to_numeric(comprehensive_df['Fixed_deferred'], errors='coerce') *
                pd.to_numeric(comprehensive_df['Projected_Infl_Factors'], errors='coerce')
            )
        else:
            comprehensive_df['Deferred_w_Real'] = None

        if 'Fixed_pensioner' in comprehensive_df.columns:
            comprehensive_df['Pensioner_w_Real'] = (
                pd.to_numeric(comprehensive_df['Fixed_pensioner'], errors='coerce') *
                pd.to_numeric(comprehensive_df['Projected_Infl_Factors'], errors='coerce')
            )
        else:
            comprehensive_df['Pensioner_w_Real'] = None
    else:
        comprehensive_df['Deferred_w_Real'] = None
        comprehensive_df['Pensioner_w_Real'] = None

    # Select base columns
    base_columns = [col for col in OUTPUT_COLUMNS_ORDER if col in comprehensive_df.columns]
    comprehensive_final = comprehensive_df[base_columns].copy()

    if 'Date_Normalized' in comprehensive_df.columns:
        comprehensive_final['Date'] = comprehensive_df['Date_Normalized']

    # Compute factors
    factors = compute_experience_and_adjustment_factors(comprehensive_final, valuation_date)
    comprehensive_final['Experience_Factor'] = factors['Experience_Factor']
    comprehensive_final['Interpolation_Vector'] = factors['Interpolation_Vector']
    comprehensive_final['Adjustment_Factor'] = factors['Adjustment_Factor']
    comprehensive_final['Credibility_Factor'] = factors['Credibility_Factor']
    comprehensive_final['X_Actual'] = factors['X_Actual']

    # Float vector calculations
    fixed_orig = pd.to_numeric(comprehensive_final['Fixed_w_Original'], errors='coerce')
    increased_orig = pd.to_numeric(comprehensive_final['Increased_w_Original'], errors='coerce')
    decreased_orig = pd.to_numeric(comprehensive_final['Decreased_w_Original'], errors='coerce')

    experience_factor = factors['Experience_Factor']
    interpolation_vector = factors['Interpolation_Vector']
    adjustment_factor = factors['Adjustment_Factor']

    mix_vector = decreased_orig if interpolation_vector == 0 else increased_orig

    comprehensive_final['Float_Vector'] = (
        (1.0 - experience_factor) * fixed_orig +
        experience_factor * mix_vector
    ) * adjustment_factor

    hist_factors = pd.to_numeric(comprehensive_final['Historical_Infl_Factors'], errors='coerce')
    proj_factors = pd.to_numeric(comprehensive_final['Projected_Infl_Factors'], errors='coerce')

    inflation_ratio = proj_factors / hist_factors.replace({0: pd.NA})
    comprehensive_final['Adjusted_Float_Vector'] = comprehensive_final['Float_Vector'] * inflation_ratio

    # Reorder columns
    ordered_cols = [col for col in OUTPUT_COLUMNS_ORDER if col in comprehensive_final.columns]
    remaining_cols = [col for col in comprehensive_final.columns if col not in ordered_cols]
    comprehensive_final = comprehensive_final[ordered_cols + remaining_cols]

    elapsed = time.time() - start_time
    logger.debug(f"Tranche {rga_tab_name} processed in {elapsed:.3f}s: {len(comprehensive_final)} rows")

    return comprehensive_final


@log_step("Processing all tranches")
def calculate_all_tranches(fixed_vectors_data, client_claims_data, client_vectors_data,
                          fee_vectors_data, valuation_date):
    """Process all tranches and return comprehensive results."""
    total_tranches = len(fixed_vectors_data)
    logger.info(f"Processing {total_tranches} tranches...")

    comprehensive_results = {}
    successful = 0
    failed = 0

    # Use progress tracker for visibility
    tracker = ProgressTracker(
        total=total_tranches,
        name="Tranche Processing",
        log_interval=20,
        logger=logger
    )

    for rga_tab_name, fixed_df in fixed_vectors_data.items():
        result = process_single_tranche(
            rga_tab_name, fixed_df, client_vectors_data, client_claims_data, valuation_date
        )

        if result is not None:
            comprehensive_results[rga_tab_name] = result
            successful += 1
            tracker.update(message=f"{rga_tab_name}: {len(result)} rows")
            logger.info(f"  ✓ {rga_tab_name}: {len(result)} rows processed")
        else:
            failed += 1
            tracker.update(message=f"{rga_tab_name}: No match")
            logger.warning(f"  ✗ {rga_tab_name}: No match found")

    tracker.complete()
    
    logger.info(f"Tranche processing complete: {successful} successful, {failed} failed out of {total_tranches}")
    log_calculation_result(logger, "Successful Tranches", successful)
    log_calculation_result(logger, "Failed Tranches", failed)
    
    # Return both comprehensive results and fee vectors data
    return comprehensive_results, fee_vectors_data


# ------------------------ Sensitivity Analysis Integration ------------------------ #


@log_step("Running sensitivity scenarios")
def run_sensitivity_runs(comprehensive_results, valuation_date, prophet_file_path, sonia_rates,
                         fee_vectors_data=None):
    """
    Run sensitivity scenarios using the new specification-compliant implementation.
    
    This is a wrapper that calls the sensitivity module's run_sensitivity_analysis function
    which properly implements:
    1. Prophet claims adjustment by inflation (Section 4.1)
    2. Blended claims calculation with BE/Stress (Section 4.2)
    3. Tranche allocation using prior month multipliers (Section 4.3)
    4. PV calculations using Prophet-based Adjusted_Float_Vector (Section 5.1)
    
    Args:
        comprehensive_results: Dict of {tranche_name: DataFrame}
        valuation_date: Valuation date string (YYYY-MM-DD)
        prophet_file_path: Path to Prophet output Excel file
        sonia_rates: Dict of {month: sonia_rate}
        fee_vectors_data: Dict of {deal_name: {'Am': DataFrame, 'In': DataFrame}} (optional)
    
    Returns:
        List of result dictionaries
    """
    from sensitivity import run_sensitivity_analysis
    
    logger.info("Delegating to specification-compliant sensitivity analysis...")
    
    # Package client data for the new function
    client_data = {
        'comprehensive_results': comprehensive_results,
        'fee_vectors_data': fee_vectors_data or {},
    }
    
    # Run the new implementation
    results_df = run_sensitivity_analysis(
        client_data=client_data,
        prophet_file_path=prophet_file_path,
        valuation_date=valuation_date,
        sonia_rates=sonia_rates,
        num_offsets=SENSITIVITY_MAX_MONTHS + 1
    )
    
    # Convert DataFrame to list of dicts for compatibility
    results = []
    for _, row in results_df.iterrows():
        results.append({
            "Deal": row["Deal Name"],
            "Scenario": row["Scenario"],
            "Date": row["Date"],
            "PV_Premium": row["PV Premium"],
            "PV_Claims": row["PV Claims"],
            "Exposure": row["Exposure"],
            "PV_Fees": row["PV Fees"],
            "Required_Amount": row["Required Amount"],
        })
    
    logger.info(f"Sensitivity calculations complete: {len(results):,} result rows generated")
    return results
