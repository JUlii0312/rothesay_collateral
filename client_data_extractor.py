"""
CLIENT DATA EXTRACTOR
=====================
Functions for extracting data from client Excel files.
"""


import pandas as pd
import os
import time
from openpyxl import load_workbook
import traceback


from logging_config import (
   get_logger,
   log_step,
   log_file_operation,
   log_calculation_result,
   log_dataframe_info,
   ProgressTracker,
)
from const import RGA_TO_CLIENT_MAPPING, get_rga_share, get_deal_name_from_rga_tab


# Initialize logger for this module
logger = get_logger("client_data_extractor")




def find_cell_position(worksheet, search_text):
   """Find the position (row, col) of a cell containing specific text."""
   logger.debug(f"Searching for cell containing: '{search_text}'")
   for row in range(1, worksheet.max_row + 1):
       for col in range(1, worksheet.max_column + 1):
           cell_value = worksheet.cell(row=row, column=col).value
           if cell_value and str(cell_value).strip().lower() == search_text.lower():
               logger.debug(f"Found '{search_text}' at row {row}, column {col}")
               return row, col
   logger.debug(f"Cell containing '{search_text}' not found")
   return None, None




def normalize_date_for_matching(date_val):
   """Normalize date values for matching between datasets."""
   if pd.isna(date_val):
       return None


   parsed_date = None
   if isinstance(date_val, pd.Timestamp):
       parsed_date = date_val
   else:
       try:
           parsed_date = pd.to_datetime(date_val, errors='coerce')
       except Exception:
           parsed_date = pd.NaT


   if parsed_date is None or pd.isna(parsed_date):
       return str(date_val)


   month_end_date = parsed_date.to_period('M').to_timestamp('M')
   return month_end_date.strftime('%Y-%m-%d')




def extract_claims_data_from_tab(worksheet, client_tab_name):
   """Extract Claims data from a single worksheet tab."""
   logger.debug(f"Extracting claims data from tab: {client_tab_name}")
   try:
       actual_claims_row, actual_claims_col = find_cell_position(worksheet, "Actual Claims")
       if actual_claims_row is None:
           logger.debug(f"'Actual Claims' section not found in tab: {client_tab_name}")
           return None


       header_row = actual_claims_row + 1


       vectors_row = None
       for row in range(header_row + 1, worksheet.max_row + 1):
           cell_value = worksheet.cell(row=row, column=1).value
           if cell_value and str(cell_value).strip().lower() == "vectors":
               vectors_row = row
               break


       if vectors_row is None:
           vectors_row = worksheet.max_row + 1


       claims_data = []
       data_start_row = header_row + 1


       for row in range(data_start_row, vectors_row):
           date_val = worksheet.cell(row=row, column=1).value
           claims_val = worksheet.cell(row=row, column=2).value


           if date_val is None or str(date_val).strip() == "":
               break


           claims_data.append({
               'Date': date_val,
               'Total_Actual_Claims': claims_val
           })


       if claims_data:
           logger.debug(f"Extracted {len(claims_data)} claims records from {client_tab_name}")
           return pd.DataFrame(claims_data)
       else:
           logger.debug(f"No claims data found in {client_tab_name}")
           return None


   except Exception as e:
       logger.error(f"Claims extraction failed for {client_tab_name}: {str(e)}")
       return None




def extract_vectors_data_from_tab(worksheet, client_tab_name):
   """Extract Vectors data from a single worksheet tab."""
   logger.debug(f"Extracting vectors data from tab: {client_tab_name}")
   try:
       vectors_row, vectors_col = find_cell_position(worksheet, "Vectors")
       if vectors_row is None:
           logger.debug(f"'Vectors' section not found in tab: {client_tab_name}")
           return None


       # Header might be merged over multiple rows; pick the first non-empty row within next 3 rows
       header_row = vectors_row + 1
       for hdr in range(vectors_row + 1, vectors_row + 4):
           row_vals = [worksheet.cell(row=hdr, column=col).value for col in range(1, worksheet.max_column + 1)]
           if any(v is not None for v in row_vals):
               header_row = hdr
               break


       column_mapping = {}
       for col in range(1, worksheet.max_column + 1):
           header_val = worksheet.cell(row=header_row, column=col).value
           if header_val:
               header_str = str(header_val).strip()
               if header_str.lower() == "date":
                   column_mapping['Date'] = col
               elif "historical infl factors" in header_str.lower():
                   column_mapping['Historical_Infl_Factors'] = col
               elif "projected infl factors" in header_str.lower():
                   column_mapping['Projected_Infl_Factors'] = col
               elif "discount factors for a and e" in header_str.lower():
                   column_mapping['Discount_Factors_A_E'] = col


       logger.debug(f"Column mapping for {client_tab_name}: {column_mapping}")


       vectors_data = []
       data_start_row = header_row + 1


       for row in range(data_start_row, worksheet.max_row + 1):
           date_val = None
           if 'Date' in column_mapping:
               date_val = worksheet.cell(row=row, column=column_mapping['Date']).value


           if date_val is None:
               break


           row_data = {'Date': date_val}
           for col_name, col_num in column_mapping.items():
               if col_name != 'Date':
                   row_data[col_name] = worksheet.cell(row=row, column=col_num).value


           vectors_data.append(row_data)


       if vectors_data:
           logger.debug(f"Extracted {len(vectors_data)} vector records from {client_tab_name}")
           return pd.DataFrame(vectors_data)
       else:
           logger.debug(f"No vectors data found in {client_tab_name}")
           return None


   except Exception as e:
       logger.error(f"Vectors extraction failed for {client_tab_name}: {str(e)}")
       return None




def combine_multiple_tabs_data(data_list, rga_tab_name, data_type="claims"):
   """Combine data from multiple client tabs."""
   logger.debug(f"Combining {len(data_list)} dataframes for {rga_tab_name} (type: {data_type})")
  
   if not data_list:
       return None


   combined_df = pd.concat(data_list, ignore_index=True)


   if data_type == "claims":
       combined_df = combined_df.groupby('Date').agg({
           'Total_Actual_Claims': 'sum'
       }).reset_index()


       rga_share = get_rga_share(rga_tab_name)
       if rga_share is not None:
           logger.debug(f"Applying RGA share {rga_share} to claims for {rga_tab_name}")
           combined_df['Total_Actual_Claims'] = (
               pd.to_numeric(combined_df['Total_Actual_Claims'], errors='coerce') / rga_share
           )


   elif data_type == "vectors":
       combined_df = combined_df.drop_duplicates(subset=['Date'], keep='first')


   logger.debug(f"Combined data has {len(combined_df)} rows for {rga_tab_name}")
   return combined_df




@log_step("Extracting client data")
def extract_client_data(client_file_path):
   """Extract Claims and Vectors data directly from client file."""
   logger.info(f"Extracting client data from: {client_file_path}")


   if not os.path.exists(client_file_path):
       logger.error(f"Client file not accessible: {client_file_path}")
       return {}, {}


   try:
       start_time = time.time()
       wb = load_workbook(client_file_path, data_only=True)
       logger.info(f"Workbook loaded: {len(wb.sheetnames)} sheets found")
       logger.debug(f"Available sheets: {wb.sheetnames[:10]}{'...' if len(wb.sheetnames) > 10 else ''}")


       client_claims_data = {}
       client_vectors_data = {}


       total_mappings = len(RGA_TO_CLIENT_MAPPING)
       processed = 0
      
       tracker = ProgressTracker(
           total=total_mappings,
           name="Client Tab Extraction",
           log_interval=25,
           logger=logger
       )


       for rga_tab, client_tab_info in RGA_TO_CLIENT_MAPPING.items():
           if '+' in client_tab_info:
               client_tabs = [tab.strip() for tab in client_tab_info.split('+')]
           else:
               client_tabs = [client_tab_info.strip()]


           claims_data_list = []
           vectors_data_list = []
           tabs_found = 0


           for client_tab in client_tabs:
               if client_tab in wb.sheetnames:
                   tabs_found += 1
                   ws = wb[client_tab]


                   claims_df = extract_claims_data_from_tab(ws, client_tab)
                   if claims_df is not None and not claims_df.empty:
                       claims_data_list.append(claims_df)


                   vectors_df = extract_vectors_data_from_tab(ws, client_tab)
                   if vectors_df is not None and not vectors_df.empty:
                       vectors_data_list.append(vectors_df)
               else:
                   logger.debug(f"Tab '{client_tab}' not found in workbook (mapped from {rga_tab})")


           if claims_data_list:
               combined_claims = combine_multiple_tabs_data(claims_data_list, rga_tab, "claims")
               if combined_claims is not None:
                   client_claims_data[rga_tab] = combined_claims
                   logger.debug(f"Claims data combined for {rga_tab}: {len(combined_claims)} rows")


           if vectors_data_list:
               combined_vectors = combine_multiple_tabs_data(vectors_data_list, rga_tab, "vectors")
               if combined_vectors is not None:
                   client_vectors_data[rga_tab] = combined_vectors
                   logger.debug(f"Vectors data combined for {rga_tab}: {len(combined_vectors)} rows")


           processed += 1
           tracker.update()


       tracker.complete()
       wb.close()
      
       elapsed = time.time() - start_time
       logger.info(f"Client data extraction completed in {elapsed:.2f}s")
       log_calculation_result(logger, "Claims Datasets Extracted", len(client_claims_data))
       log_calculation_result(logger, "Vectors Datasets Extracted", len(client_vectors_data))
      
       return client_claims_data, client_vectors_data


   except Exception as e:
       logger.exception(f"Error extracting client data: {str(e)}")
       return {}, {}




@log_step("Extracting client discount factors")
def extract_client_discount_factors(client_file_path):
   """Extract Client Discount Factors from client file."""
   logger.info(f"Extracting discount factors from: {client_file_path}")


   if not os.path.exists(client_file_path):
       logger.error(f"Client file not accessible: {client_file_path}")
       return None


   try:
       wb = load_workbook(client_file_path, data_only=True)
       tabs_searched = 0


       for rga_tab, client_tab_info in RGA_TO_CLIENT_MAPPING.items():
           if '+' in client_tab_info:
               client_tabs = [tab.strip() for tab in client_tab_info.split('+')]
           else:
               client_tabs = [client_tab_info.strip()]


           for client_tab in client_tabs:
               if client_tab in wb.sheetnames:
                   tabs_searched += 1
                   ws = wb[client_tab]


                   vectors_row, _ = find_cell_position(ws, "Vectors")
                   if vectors_row is None:
                       continue


                   date_col = None
                   discount_col = None


                   # Try up to 3 rows for header (to tolerate merged/blank spacer rows)
                   for hdr in range(vectors_row + 1, vectors_row + 4):
                       for col in range(1, ws.max_column + 1):
                           header_val = ws.cell(row=hdr, column=col).value
                           if header_val:
                               header_str = str(header_val).strip().lower()
                               if header_str == "date":
                                   date_col = col
                               elif "discount factors for a and e" in header_str:
                                   discount_col = col
                       if date_col is not None and discount_col is not None:
                           header_row = hdr
                           break


                   if date_col is None or discount_col is None:
                       continue


                   logger.debug(f"Found discount factors in {client_tab} at row {header_row}")


                   discount_data = []
                   data_start_row = header_row + 1


                   for row in range(data_start_row, ws.max_row + 1):
                       date_val = ws.cell(row=row, column=date_col).value
                       if date_val is None:
                           break


                       discount_val = ws.cell(row=row, column=discount_col).value
                       discount_data.append({
                           'Date': date_val,
                           'Client_Discount_Factor': discount_val
                       })


                   if discount_data:
                       df = pd.DataFrame(discount_data)
                       wb.close()
                       logger.info(f"Extracted {len(df)} discount factor records from {client_tab}")
                       return df


       wb.close()
       logger.warning(f"No discount factors found after searching {tabs_searched} tabs")
       return None


   except Exception as e:
       logger.exception(f"Error extracting discount factors: {str(e)}")
       return None




@log_step("Extracting client exposure summary")
def extract_client_exposure_summary(client_file_path):
   """Extract Fee PV, A PV, E PV, Experience Factor, Exposure values from client file tabs."""
   logger.info(f"Extracting exposure summary from: {client_file_path}")


   if not os.path.exists(client_file_path):
       logger.error(f"Client file not accessible: {client_file_path}")
       return []


   try:
       wb = load_workbook(client_file_path, data_only=True)
       exposure_data = []


       for rga_tab, client_tab_info in RGA_TO_CLIENT_MAPPING.items():
           deal_name = get_deal_name_from_rga_tab(rga_tab)
           if rga_tab.startswith(deal_name + "_"):
               tranche_suffix = rga_tab[len(deal_name) + 1:]
           else:
               tranche_suffix = rga_tab


           if '+' in client_tab_info:
               client_tabs = [tab.strip() for tab in client_tab_info.split('+')]
           else:
               client_tabs = [client_tab_info.strip()]


           total_values = {'Fee_PV': 0.0, 'A_PV': 0.0, 'E_PV': 0.0, 'Exposure': 0.0, 'Experience_Factor': 0.0}
           tabs_found = 0
           experience_factor_set = False


           for client_tab in client_tabs:
               if client_tab in wb.sheetnames:
                   ws = wb[client_tab]
                   tabs_found += 1


                   for row in range(1, ws.max_row + 1):
                       cell_a = ws.cell(row=row, column=1).value
                       if cell_a:
                           cell_a_lower = str(cell_a).strip().lower()
                           cell_b_value = ws.cell(row=row, column=2).value


                           try:
                               numeric_value = float(cell_b_value) if cell_b_value else 0.0
                           except (ValueError, TypeError):
                               numeric_value = 0.0


                           if cell_a_lower == 'fee pv':
                               total_values['Fee_PV'] += numeric_value
                           elif cell_a_lower == 'a pv':
                               total_values['A_PV'] += numeric_value
                           elif cell_a_lower == 'e pv':
                               total_values['E_PV'] += numeric_value
                           elif cell_a_lower == 'exposure':
                               total_values['Exposure'] += numeric_value
                           elif cell_a_lower == 'experience factor' and not experience_factor_set:
                               total_values['Experience_Factor'] = numeric_value
                               experience_factor_set = True


           if tabs_found > 0:
               exposure_data.append({
                   'Deal': deal_name,
                   'Tranche': tranche_suffix,
                   'Fee_PV': total_values['Fee_PV'],
                   'A_PV': total_values['A_PV'],
                   'E_PV': total_values['E_PV'],
                   'Exposure': total_values['Exposure'],
                   'Experience_Factor': total_values['Experience_Factor']
               })
               logger.debug(f"Extracted exposure for {deal_name}/{tranche_suffix}: "
                          f"Fee_PV={total_values['Fee_PV']:.2f}, Exposure={total_values['Exposure']:.2f}")


       wb.close()
       logger.info(f"Extracted {len(exposure_data)} exposure summary entries")
       return exposure_data


   except Exception as e:
       logger.exception(f"Error extracting exposure summary: {str(e)}")
       return []




@log_step("Extracting fee vectors")
def extract_fee_vectors(client_file_path):
   """Extract Fee Vector data from fee tabs for AdvanceT1A and AdvanceT1B.
    Returns fee vectors separately by Am and In tabs for each deal.
   Format: {deal_name: {'Am': DataFrame, 'In': DataFrame}}
   """
   from const import FEE_TAB_MAPPING
   
   logger.info(f"Extracting fee vectors from: {client_file_path}")


   if not os.path.exists(client_file_path):
       logger.error(f"Client file not accessible: {client_file_path}")
       return {}


   try:
       wb = load_workbook(client_file_path, data_only=True)
       fee_vectors_data = {}


       for deal_name, fee_tab_info in FEE_TAB_MAPPING.items():
           logger.debug(f"Processing fee vectors for deal: {deal_name}")
          
           if '+' in fee_tab_info:
               fee_tabs = [tab.strip() for tab in fee_tab_info.split('+')]
           else:
               fee_tabs = [fee_tab_info.strip()]


           fee_data_by_type = {}


           for fee_tab in fee_tabs:
               if fee_tab in wb.sheetnames:
                   ws = wb[fee_tab]


                   # Find "Vectors" section (same as discount factors extraction)
                   vectors_row, _ = find_cell_position(ws, "Vectors")
                   if vectors_row is None:
                       logger.debug(f"'Vectors' section not found in {fee_tab}")
                       continue


                   # Header may be merged; find first non-empty row within next 3 rows
                   header_row = vectors_row + 1
                   for hdr in range(vectors_row + 1, vectors_row + 4):
                       row_vals = [ws.cell(row=hdr, column=c).value for c in range(1, ws.max_column + 1)]
                       if any(v is not None for v in row_vals):
                           header_row = hdr
                           break


                   # Find column mappings in header row (same logic as Discount Factors)
                   column_mapping = {}
                   for col in range(1, ws.max_column + 1):
                       header_val = ws.cell(row=header_row, column=col).value
                       if header_val:
                           header_str = str(header_val).strip().lower()
                           if header_str == "date":
                               column_mapping['Date'] = col
                           elif "fee vector" in header_str:
                               column_mapping['Fee_Vector'] = col


                   # Check if we found both required columns
                   if 'Date' not in column_mapping or 'Fee_Vector' not in column_mapping:
                       logger.warning(f"{fee_tab}: Missing Date or Fee Vector column")
                       continue


                   # Extract data row by row (vertical layout, same as Discount Factors)
                   fee_data = []
                   data_start_row = header_row + 1


                   for row in range(data_start_row, ws.max_row + 1):
                       date_val = ws.cell(row=row, column=column_mapping['Date']).value
                       if date_val is None:
                           break


                       fee_val = ws.cell(row=row, column=column_mapping['Fee_Vector']).value
                       fee_data.append({
                           'Date': date_val,
                           'Fee_Vector': fee_val
                       })


                   if fee_data:
                       # Determine if this is Am or In tab
                       if 'RGAAm' in fee_tab:
                           fee_type = 'Am'
                       elif 'RGAIn' in fee_tab:
                           fee_type = 'In'
                       else:
                           continue
                      
                       fee_data_by_type[fee_type] = pd.DataFrame(fee_data)
                       logger.info(f"Extracted {len(fee_data)} fee vector records from {fee_tab}")


           # Store separate Am and In data for this deal
           if fee_data_by_type:
               fee_vectors_data[deal_name] = fee_data_by_type


       wb.close()
       logger.info(f"Fee vectors extracted for {len(fee_vectors_data)} deals")
       return fee_vectors_data


   except Exception as e:
       logger.exception(f"Error extracting fee vectors: {str(e)}")
       return {}




@log_step("Extracting Fee PV values")
def extract_fee_pv_values(client_file_path):
   """Extract Fee PV values from fee tabs."""
   from const import FEE_TAB_MAPPING
   
   logger.info(f"Extracting Fee PV values from: {client_file_path}")


   if not os.path.exists(client_file_path):
       logger.error(f"Client file not accessible: {client_file_path}")
       return {}


   try:
       wb = load_workbook(client_file_path, data_only=True)
       fee_pv_data = {}


       for deal_name, fee_tab_info in FEE_TAB_MAPPING.items():
           if '+' in fee_tab_info:
               fee_tabs = [tab.strip() for tab in fee_tab_info.split('+')]
           else:
               fee_tabs = [fee_tab_info.strip()]


           total_fee_pv = 0.0


           for fee_tab in fee_tabs:
               if fee_tab in wb.sheetnames:
                   ws = wb[fee_tab]


                   # Search for "Fee PV" in column A
                   for row in range(1, ws.max_row + 1):
                       cell_a = ws.cell(row=row, column=1).value
                       if cell_a and str(cell_a).strip().lower() == "fee pv":
                           fee_pv_val = ws.cell(row=row, column=2).value
                           try:
                               total_fee_pv += float(fee_pv_val) if fee_pv_val else 0.0
                               logger.debug(f"Found Fee PV in {fee_tab}: {fee_pv_val}")
                           except (ValueError, TypeError):
                               pass
                           break


           if total_fee_pv != 0.0:
               fee_pv_data[deal_name] = total_fee_pv
               logger.debug(f"Total Fee PV for {deal_name}: {total_fee_pv:.2f}")


       wb.close()
       logger.info(f"Fee PV extracted for {len(fee_pv_data)} deals")
       return fee_pv_data


   except Exception as e:
       logger.exception(f"Error extracting Fee PV: {str(e)}")
       return {}




@log_step("Loading fixed vectors data")
def load_fixed_vectors_data(file_path):
   """Load all Fixed Vectors data from the consolidated RGA file."""
   logger.info(f"Loading fixed vectors from: {file_path}")


   if not os.path.exists(file_path):
       logger.error(f"Fixed Vectors file not found: {file_path}")
       return {}


   try:
       start_time = time.time()
       all_sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
      
       # Filter out summary sheets
       fixed_vectors_data = {
           sheet_name: df for sheet_name, df in all_sheets.items()
           if not sheet_name.endswith('_Summary') and sheet_name != 'Extraction_Summary'
       }
      
       elapsed = time.time() - start_time
       logger.info(f"Loaded {len(fixed_vectors_data)} fixed vector datasets in {elapsed:.2f}s")
      
       # Log sample info for first few datasets
       for i, (name, df) in enumerate(fixed_vectors_data.items()):
           if i < 3:  # Log first 3
               log_dataframe_info(logger, df, f"Fixed Vectors - {name}")
           elif i == 3:
               logger.debug(f"... and {len(fixed_vectors_data) - 3} more datasets")
               break
      
       return fixed_vectors_data


   except Exception as e:
       logger.exception(f"Error loading fixed vectors: {str(e)}")
       return {}



