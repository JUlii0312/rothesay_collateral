"""
SENSITIVITY ANALYSIS MODULE
===========================
Contains all sensitivity calculation logic for UK Longevity Swap Calculator.

This module implements the calculation flow per specification:
1. Load Prophet claims (BE and Stress)
2. Adjust Prophet claims by inflation (Section 4.1)
3. Blend BE and Stress claims based on grading period and offset (Section 4.2)
4. Allocate to tranches using multipliers from prior month claims (Section 4.3)
5. Calculate PV metrics using adjusted Prophet claims (Section 5.1)
"""

import pandas as pd
import numpy as np
import time
from datetime import datetime
from dateutil.relativedelta import relativedelta
from typing import Dict, Optional, Union, List, Tuple
import os

from logging_config import (
    get_logger,
    log_step,
    log_calculation_result,
    log_dict_summary,
    log_dataframe_info,
    ProgressTracker,
)

# Initialize logger for this module
logger = get_logger("sensitivity")

# =============================================================================
# Configuration Constants (Section 2.2)
# =============================================================================

SCENARIOS = [
    {
        "name": "Scenario 1",
        "claims_type": "BE",
        "grading_period": 0,
        "discount_rate_shock": 0.0
    },
    {
        "name": "Scenario 2",
        "claims_type": "Blended_1in200",
        "grading_period": 60,
        "discount_rate_shock": 0.0
    },
    {
        "name": "Scenario 3",
        "claims_type": "Blended_1in200",
        "grading_period": 60,
        "discount_rate_shock": 0.013411942
    },
    {
        "name": "Scenario 4",
        "claims_type": "Blended_1in200",
        "grading_period": 60,
        "discount_rate_shock": 0.011564121
    },
    {
        "name": "Scenario 5",
        "claims_type": "Blended_1in200",
        "grading_period": 60,
        "discount_rate_shock": -0.003345999
    },
    {
        "name": "Scenario 6",
        "claims_type": "Blended_1in200",
        "grading_period": 60,
        "discount_rate_shock": -0.009482681
    }
]

# Deal to Prophet column header mapping (Section 3.4)
DEAL_PROPHET_MAPPING = {
    "Artemis": "Rothesay",
    "Excalibur": "Rothesay_Excalibur",
    "Lancelot": "Rothesay_Lancelot",
    "Paternoster": "Paternoster",
    "Romeo": "Romeo",
    "Titan": "Titan",
    "Advance_T1a": "Rothesay_Advance",
    "Advance_T1b": "Paternoster",
    "Antigua": "Paternoster",
    "Jupiter": "Rothesay_Jupiter",
    "Laker_T1": "Paternoster",
    "Sherwood_Forest": "Paternoster"
}

# Prophet sheet names by claims type (Section 3.2)
PROPHET_SHEETS = {
    "BE": "Longevity_GFS_RM",
    "1in200": "Longevity_OVR_TOR9",
    "1in20": "Longevity_OVR_TOR3"
}

# SONIA adjustment constant (0.25%) - Section 4.4
SONIA_ADJUSTMENT = 0.0025

# Maximum months for sensitivity offsets
SENSITIVITY_MAX_MONTHS = 500


# =============================================================================
# Data Loading Functions
# =============================================================================

def _normalize_prophet_date(date_val) -> Optional[str]:
    """Normalize Prophet YYYYMM date values to month-end string."""
    if pd.isna(date_val):
        return None
    try:
        date_str = str(int(date_val))
        if len(date_str) == 6:
            # YYYYMM format
            parsed = pd.to_datetime(f"{date_str}01", format="%Y%m%d", errors="coerce")
            if pd.isna(parsed):
                return None
            return parsed.to_period("M").to_timestamp("M").strftime("%Y-%m-%d")
        # Try parsing as regular date
        parsed = pd.to_datetime(date_val, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_period("M").to_timestamp("M").strftime("%Y-%m-%d")
    except Exception:
        return None


@log_step("Loading Prophet data")
def load_prophet_data(file_path: str, tab_name: str) -> pd.DataFrame:
    """
    Load Prophet data from Excel file (Section 3.2).
    
    Args:
        file_path: Path to Prophet Excel file
        tab_name: Sheet name to load (e.g., 'Longevity_GFS_RM')
    
    Returns:
        DataFrame with Date and prophet columns
    """
    logger.info(f"Loading Prophet data from sheet: {tab_name}")
    
    if not os.path.exists(file_path):
        logger.error(f"Prophet file not found: {file_path}")
        raise FileNotFoundError(f"Prophet file not found: {file_path}")
    
    try:
        # Read raw data - Row 1 has prophet names, Column B has dates in YYYYMM format
        raw_df = pd.read_excel(file_path, sheet_name=tab_name, header=None)
        
        if raw_df.empty or raw_df.shape[1] < 3:
            logger.warning(f"Prophet sheet {tab_name} is empty or has insufficient columns")
            return pd.DataFrame()
        
        logger.debug(f"Raw data shape: {raw_df.shape}")
        
        # Extract prophet names from row 0, starting from column 2
        prophet_names = raw_df.iloc[0, 2:].tolist()
        logger.debug(f"Found {len(prophet_names)} prophet series")
        
        # Extract dates from column 1, starting from row 1
        dates = raw_df.iloc[1:, 1].tolist()
        
        # Extract data block
        data_block = raw_df.iloc[1:, 2:]
        data_block.columns = prophet_names
        data_block.insert(0, 'Date', dates)
        
        # Normalize dates to month-end format
        data_block['Date'] = data_block['Date'].apply(_normalize_prophet_date)
        data_block = data_block.dropna(subset=['Date'])
        
        logger.info(f"Loaded Prophet data: {len(data_block)} rows, {len(prophet_names)} series")
        
        return data_block
        
    except Exception as e:
        logger.exception(f"Error loading Prophet data from {tab_name}: {str(e)}")
        raise ValueError(f"Error loading Prophet data from {tab_name}: {str(e)}")


def get_prophet_claims_for_deal(prophet_df: pd.DataFrame, deal_name: str) -> pd.Series:
    """
    Extract claims cashflow for a specific deal from Prophet data.
    
    Args:
        prophet_df: DataFrame loaded from Prophet file
        deal_name: Deal name (e.g., 'Artemis')
    
    Returns:
        Series of claims indexed by date
    """
    logger.debug(f"Extracting Prophet claims for deal: {deal_name}")
    
    prophet_column = DEAL_PROPHET_MAPPING.get(deal_name)
    if prophet_column is None:
        logger.error(f"Unknown deal: {deal_name}")
        raise ValueError(f"Unknown deal: {deal_name}")
    
    if prophet_column not in prophet_df.columns:
        logger.error(f"Prophet column '{prophet_column}' not found in data")
        raise ValueError(f"Prophet column '{prophet_column}' not found in data")
    
    result = prophet_df[['Date', prophet_column]].copy()
    result = result.set_index('Date')[prophet_column]
    result = pd.to_numeric(result, errors='coerce').fillna(0)
    
    logger.debug(f"Extracted {len(result)} claims values for {deal_name}")
    return result


# =============================================================================
# Claims Adjustment Functions (Section 4.1)
# =============================================================================

def adjust_prophet_claims(
    prophet_claims: pd.Series,
    rga_share: float,
    projected_infl_factors: Dict[str, float],
    valdate_infl_factor: float
) -> pd.Series:
    """
    DEPRECATED: This function adjusts claims at deal level before tranche allocation.
    
    The new approach applies:
    1. RGA share at deal level (in Step 2)
    2. Inflation adjustment at tranche level (in precompute_blending_vectors)
    
    This function is kept for reference but should not be used.
    
    Original formula:
        adjusted_claims = prophet_output / rga_share * projected_infl_factors / valdate_infl_factor
    
    Args:
        prophet_claims: Raw claims from Prophet output (indexed by date)
        rga_share: RGA share percentage for the deal
        projected_infl_factors: Dict of {date_str: inflation_factor}
        valdate_infl_factor: Projected inflation factor at valuation date
    
    Returns:
        Adjusted claims series
    """
    import warnings
    warnings.warn(
        "adjust_prophet_claims is deprecated. Inflation is now applied per-tranche in precompute_blending_vectors.",
        DeprecationWarning,
        stacklevel=2
    )
    logger.debug(f"Adjusting Prophet claims: RGA share={rga_share}, valdate_infl={valdate_infl_factor}")
    
    if rga_share == 0:
        logger.error("RGA share cannot be zero")
        raise ValueError("RGA share cannot be zero")
    if valdate_infl_factor == 0 or pd.isna(valdate_infl_factor):
        logger.warning("Valuation date inflation factor is zero/NA, using 1.0")
        valdate_infl_factor = 1.0
    
    adjusted = prophet_claims.copy()
    
    for date_str in adjusted.index:
        infl_factor = projected_infl_factors.get(date_str, valdate_infl_factor)
        if pd.isna(infl_factor) or infl_factor == 0:
            infl_factor = valdate_infl_factor
        
        # Formula: prophet_output / rga_share * projected_infl / valdate_infl
        adjusted[date_str] = (prophet_claims[date_str] / rga_share) * (infl_factor / valdate_infl_factor)
    
    return adjusted


# =============================================================================
# Blended Claims Calculation (Section 4.2)
# =============================================================================

def calculate_blended_claims(
    be_claims: pd.Series,
    stress_claims: pd.Series,
    grading_period: int,
    offset: int
) -> pd.Series:
    """
    Calculate blended claims using linear interpolation (Section 4.2).
    
    Formula:
        if grading_period == 0: return be_claims
        blend_factor = min(1.0, offset / grading_period)
        blended = (1 - blend_factor) * be_claims + blend_factor * stress_claims
    
    Args:
        be_claims: Best Estimate claims cashflow (adjusted)
        stress_claims: 1-in-200 (or 1-in-20) stress claims cashflow (adjusted)
        grading_period: Number of months for full transition (e.g., 60)
        offset: Current sensitivity offset (0 to 500)
    
    Returns:
        Blended claims series
    
    Key Logic:
        - When offset = 0: Use 100% BE claims
        - When offset = grading_period (60): Use 100% stress claims
        - When offset > grading_period: Use 100% stress claims
    """
    if grading_period == 0:
        return be_claims.copy()
    
    blend_factor = min(1.0, offset / grading_period)
    
    # Align indices
    common_dates = be_claims.index.intersection(stress_claims.index)
    be_aligned = be_claims.reindex(common_dates).fillna(0)
    stress_aligned = stress_claims.reindex(common_dates).fillna(0)
    
    blended = (1 - blend_factor) * be_aligned + blend_factor * stress_aligned
    
    return blended


# =============================================================================
# Tranche Allocation (Section 4.3)
# =============================================================================

def calculate_tranche_multipliers(
    comprehensive_results: Dict[str, pd.DataFrame],
    valuation_date: str,
    deal_name: str
) -> Dict[str, float]:
    """
    Calculate allocation multipliers for each tranche within a deal (Section 4.3).
    
    Uses Total_Actual_Claims from one month prior to valuation date.
    
    Args:
        comprehensive_results: Dict of {tranche_name: DataFrame with Total_Actual_Claims}
        valuation_date: Current valuation date (YYYY-MM-DD format)
        deal_name: Deal name to filter tranches
    
    Returns:
        Dict of {tranche_name: multiplier}
    """
    from const import get_deal_name_from_rga_tab
    
    logger.debug(f"Calculating tranche multipliers for deal {deal_name}, valuation date: {valuation_date}")
    
    # Get prior month date
    val_date = pd.to_datetime(valuation_date)
    prior_month = (val_date - relativedelta(months=1)).to_period('M').to_timestamp('M')
    prior_month_str = prior_month.strftime("%Y-%m-%d")
    
    # Get tranches for this deal
    deal_tranches = {
        name: df for name, df in comprehensive_results.items()
        if get_deal_name_from_rga_tab(name) == deal_name
    }
    
    if not deal_tranches:
        logger.warning(f"No tranches found for deal: {deal_name}")
        return {}
    
    # Get claims at prior month for each tranche
    tranche_claims = {}
    for tranche_name, df in deal_tranches.items():
        if df is None or df.empty:
            tranche_claims[tranche_name] = 0.0
            continue
        
        df_local = df.copy()
        df_local['Date_Datetime'] = pd.to_datetime(df_local['Date'], errors='coerce')
        
        # Find row for prior month
        mask = df_local['Date_Datetime'] == prior_month
        if mask.any():
            claim_val = pd.to_numeric(
                df_local.loc[mask, 'Total_Actual_Claims'].iloc[0],
                errors='coerce'
            )
            tranche_claims[tranche_name] = float(claim_val) if not pd.isna(claim_val) else 0.0
        else:
            # Try string matching
            if prior_month_str in df_local['Date'].values:
                idx = df_local[df_local['Date'] == prior_month_str].index[0]
                claim_val = pd.to_numeric(
                    df_local.loc[idx, 'Total_Actual_Claims'],
                    errors='coerce'
                )
                tranche_claims[tranche_name] = float(claim_val) if not pd.isna(claim_val) else 0.0
            else:
                tranche_claims[tranche_name] = 0.0
    
    total_claims = sum(tranche_claims.values())
    
    # Calculate multipliers
    if total_claims == 0:
        # Equal distribution if no prior claims
        n_tranches = len(tranche_claims)
        logger.debug(f"No prior claims - using equal distribution across {n_tranches} tranches")
        return {tranche: 1.0 / n_tranches for tranche in tranche_claims}
    
    multipliers = {
        tranche: claims / total_claims
        for tranche, claims in tranche_claims.items()
    }
    
    logger.debug(f"Calculated multipliers for {len(multipliers)} tranches, sum={sum(multipliers.values()):.6f}")
    return multipliers


def allocate_claims_to_tranches(
    deal_claims: pd.Series,
    multipliers: Dict[str, float]
) -> Dict[str, pd.Series]:
    """
    Apply multipliers to distribute deal-level Prophet cashflows to tranches (Section 4.3).
    
    Formula:
        tranche_prophet_cf = deal_prophet_cf * tranche_multiplier
    
    Args:
        deal_claims: Deal-level adjusted claims series
        multipliers: Dict of {tranche_name: multiplier}
    
    Returns:
        Dict of {tranche_name: allocated_claims_series}
    """
    allocated = {}
    for tranche_name, multiplier in multipliers.items():
        allocated[tranche_name] = deal_claims * multiplier
    return allocated


# =============================================================================
# Discount Rate Calculation Functions (Section 4.4)
# =============================================================================

def calculate_scenario_discount_curve(
    base_sonia: Dict[int, float],
    discount_shock: float,
    max_months: int = 601
) -> pd.DataFrame:
    """
    Build discount curve with shock applied (Section 4.4).
    
    Formula:
        adjusted_sonia = sonia_rate + discount_shock + SONIA_ADJUSTMENT
        monthly_rate = (1 + adjusted_sonia)^(1/12) - 1
        discount_factor = 1 / (1 + monthly_rate)^month_number
    
    Args:
        base_sonia: Dict of {month: sonia_rate}
        discount_shock: Discount rate shock to apply
        max_months: Maximum number of months
    
    Returns:
        DataFrame with Month, Date placeholder, Final_Discount_Curve
    """
    logger.debug(f"Building scenario discount curve: shock={discount_shock:.6f}, max_months={max_months}")
    
    discount_curve_data = []
    
    for month in range(max_months):
        if month == 0:
            discount_curve_data.append({
                'Month': month,
                'Final_Discount_Curve': 1.0
            })
            continue
        
        # Get SONIA rate for this month (use closest available if exact not found)
        sonia_rate = base_sonia.get(month)
        if sonia_rate is None:
            available_months = [m for m in base_sonia.keys() if m <= month]
            if available_months:
                sonia_rate = base_sonia[max(available_months)]
            else:
                sonia_rate = 0.04  # Default fallback
        
        # Apply shock and SONIA adjustment (Section 4.4)
        adjusted_sonia = sonia_rate + discount_shock + SONIA_ADJUSTMENT
        
        # Convert to monthly rate
        monthly_rate = (1 + adjusted_sonia) ** (1/12) - 1
        
        # Calculate discount factor
        discount_factor = 1 / ((1 + monthly_rate) ** month)
        discount_curve_data.append({
            'Month': month,
            'Final_Discount_Curve': discount_factor
        })
    
    result_df = pd.DataFrame(discount_curve_data)
    logger.debug(f"Built discount curve: {len(result_df)} factors")
    return result_df


def shift_discount_curve(
    base_curve_df: pd.DataFrame,
    offset: int
) -> np.ndarray:
    """
    Parallel shift discount curve forward by offset months.
    
    The curve shape is preserved - same factor values apply to each
    month-to-month transition regardless of starting point.
    
    Args:
        base_curve_df: Original discount curve DataFrame
        offset: Number of months to shift forward
    
    Returns:
        NumPy array of discount factors (truncated by offset length)
    """
    factors = base_curve_df['Final_Discount_Curve'].values
    
    if offset >= len(factors):
        return np.array([1.0])
    
    # Parallel shift: use same curve, just truncated
    max_len = len(factors) - offset
    return factors[:max_len].copy()


def precompute_shifted_curves(
    base_curve_df: pd.DataFrame,
    num_offsets: int
) -> List[np.ndarray]:
    """
    Pre-compute all shifted discount curves for efficient offset lookup.
    
    Instead of calling shift_discount_curve() for each offset, this function
    pre-computes all shifted curves once and returns them as a list.
    
    Args:
        base_curve_df: Original discount curve DataFrame
        num_offsets: Number of offsets to pre-compute (0 to num_offsets-1)
    
    Returns:
        List of NumPy arrays, where index i contains the curve shifted by i months
    """
    factors = base_curve_df['Final_Discount_Curve'].values
    curve_len = len(factors)
    
    shifted_curves = []
    for offset in range(num_offsets):
        if offset >= curve_len:
            shifted_curves.append(np.array([1.0]))
        else:
            max_len = curve_len - offset
            shifted_curves.append(factors[:max_len].copy())
    
    return shifted_curves


def build_discount_curve_vectorized(
    base_sonia: Dict[int, float],
    discount_shock: float,
    max_months: int = 601
) -> np.ndarray:
    """
    Build discount curve with shock applied using vectorized operations.
    
    This is a faster alternative to calculate_scenario_discount_curve() that
    returns a NumPy array directly instead of a DataFrame.
    
    Args:
        base_sonia: Dict of {month: sonia_rate}
        discount_shock: Discount rate shock to apply
        max_months: Maximum number of months
    
    Returns:
        NumPy array of discount factors
    """
    # Build SONIA rates array
    sonia_array = np.zeros(max_months)
    sonia_array[0] = 0.0  # Month 0 doesn't need SONIA
    
    # Get sorted available months for interpolation
    if base_sonia:
        sorted_months = sorted(base_sonia.keys())
        
        for month in range(1, max_months):
            rate = base_sonia.get(month)
            if rate is None:
                # Find closest available month
                available = [m for m in sorted_months if m <= month]
                if available:
                    rate = base_sonia[max(available)]
                else:
                    rate = 0.04  # Default fallback
            sonia_array[month] = rate
    else:
        sonia_array[1:] = 0.04  # Default rate
    
    # Calculate discount factors vectorized
    discount_factors = np.ones(max_months)
    
    # adjusted_sonia = sonia_rate + discount_shock + SONIA_ADJUSTMENT
    adjusted_sonia = sonia_array[1:] + discount_shock + SONIA_ADJUSTMENT
    
    # monthly_rate = (1 + adjusted_sonia)^(1/12) - 1
    monthly_rates = np.power(1 + adjusted_sonia, 1/12) - 1
    
    # discount_factor = 1 / (1 + monthly_rate)^month
    months = np.arange(1, max_months)
    discount_factors[1:] = 1.0 / np.power(1 + monthly_rates, months)
    
    return discount_factors


# =============================================================================
# PV Calculation Functions (Section 5.1)
# =============================================================================

def calculate_pv_premium(
    fixed_w_real: np.ndarray,
    discount_curve: np.ndarray,
    rga_share: float
) -> float:
    """
    Calculate Present Value of Premium (Section 5.1).
    
    Formula:
        PV Premium = SUMPRODUCT(Fixed_w_Real, Discount_Curve) * RGA_Share
    
    Args:
        fixed_w_real: Fixed vectors with real inflation
        discount_curve: Discount factors
        rga_share: RGA share percentage
    
    Returns:
        PV of Premium
    """
    # Ensure arrays are same length
    min_len = min(len(fixed_w_real), len(discount_curve))
    fixed_w_real = np.asarray(fixed_w_real[:min_len])
    discount_curve = np.asarray(discount_curve[:min_len])
    
    return np.nansum(fixed_w_real * discount_curve) * rga_share


def calculate_pv_claims(
    adjusted_float_vector: np.ndarray,
    discount_curve: np.ndarray,
    rga_share: float
) -> float:
    """
    Calculate Present Value of Claims (Section 5.1).
    
    Formula:
        PV Claims = SUMPRODUCT(Adjusted_Float_Vector, Discount_Curve) * RGA_Share
    
    Note: The Adjusted_Float_Vector here is the Prophet-based adjusted claims
    that have been blended and allocated to tranches.
    
    Args:
        adjusted_float_vector: Adjusted Prophet claims (blended, allocated)
        discount_curve: Discount factors
        rga_share: RGA share percentage
    
    Returns:
        PV of Claims
    """
    # Ensure arrays are same length
    min_len = min(len(adjusted_float_vector), len(discount_curve))
    adjusted_float_vector = np.asarray(adjusted_float_vector[:min_len])
    discount_curve = np.asarray(discount_curve[:min_len])
    
    return np.nansum(adjusted_float_vector * discount_curve) * rga_share


def calculate_exposure(pv_claims: float, pv_premium: float) -> float:
    """
    Calculate Exposure (Section 5.1).
    
    Formula:
        Exposure = PV Claims - PV Premium + 1% * PV Premium
    
    Args:
        pv_claims: Present value of claims
        pv_premium: Present value of premium
    
    Returns:
        Exposure value
    """
    return pv_claims - pv_premium + 0.01 * pv_premium


def calculate_pv_fees(
    fixed_w_real: Optional[np.ndarray],
    discount_curve: np.ndarray,
    rga_share: float,
    gross_up_factor: float = 0.0,
    is_deferred_pensioner: bool = False,
    deferred_w_real: Optional[np.ndarray] = None,
    pensioner_w_real: Optional[np.ndarray] = None,
    gross_up_deferred: float = 0.0,
    gross_up_pensioner: float = 0.0
) -> float:
    """
    Calculate Present Value of Fees (Section 5.2).
    
    Formula:
        For regular tranches:
            fee_cashflow = Fixed_w_Real * Gross_Up_Factor * RGA_Share
        For deferred/pensioner tranches:
            fee_cashflow = (Deferred_w_Real * Gross_Up_Deferred + 
                           Pensioner_w_Real * Gross_Up_Pensioner) * RGA_Share
        PV Fees = SUMPRODUCT(fee_cashflow, scenario_discount_curve)
    
    Args:
        fixed_w_real: Fixed vectors with real inflation (for regular tranches)
        discount_curve: Scenario discount factors
        rga_share: RGA share percentage
        gross_up_factor: Standard gross-up factor (for regular tranches)
        is_deferred_pensioner: Whether this is a deferred/pensioner tranche
        deferred_w_real: Deferred vectors (for deferred/pensioner tranches)
        pensioner_w_real: Pensioner vectors (for deferred/pensioner tranches)
        gross_up_deferred: Deferred gross-up factor
        gross_up_pensioner: Pensioner gross-up factor
    
    Returns:
        PV of Fees
    """
    discount_curve = np.asarray(discount_curve)
    
    if is_deferred_pensioner:
        if deferred_w_real is None or pensioner_w_real is None:
            return 0.0
        
        deferred_w_real = np.asarray(deferred_w_real)
        pensioner_w_real = np.asarray(pensioner_w_real)
        
        # Ensure arrays are same length
        min_len = min(len(deferred_w_real), len(pensioner_w_real), len(discount_curve))
        deferred_w_real = deferred_w_real[:min_len]
        pensioner_w_real = pensioner_w_real[:min_len]
        discount_curve = discount_curve[:min_len]
        
        fee_cashflow = (deferred_w_real * gross_up_deferred + 
                       pensioner_w_real * gross_up_pensioner) * rga_share
    else:
        if fixed_w_real is None:
            return 0.0
        
        fixed_w_real = np.asarray(fixed_w_real)
        
        # Ensure arrays are same length
        min_len = min(len(fixed_w_real), len(discount_curve))
        fixed_w_real = fixed_w_real[:min_len]
        discount_curve = discount_curve[:min_len]
        
        fee_cashflow = fixed_w_real * gross_up_factor * rga_share
    
    return -np.nansum(fee_cashflow * discount_curve)


def calculate_required_amount(exposure: float) -> float:
    """
    Calculate Required Amount (Section 5.1).
    
    Formula:
        Required Amount = Exposure (direct passthrough)
    
    Args:
        exposure: Calculated exposure
    
    Returns:
        Required amount (same as exposure)
    """
    return exposure


# =============================================================================
# Helper Functions
# =============================================================================

def get_prior_month(valuation_date: str) -> str:
    """Get the prior month date string."""
    val_date = pd.to_datetime(valuation_date)
    prior_month = (val_date - relativedelta(months=1)).to_period('M').to_timestamp('M')
    return prior_month.strftime("%Y-%m-%d")


def get_offset_date(valuation_date: str, offset: int) -> str:
    """Get the date for a given offset from valuation date."""
    val_date = pd.to_datetime(valuation_date)
    offset_date = (val_date + relativedelta(months=offset)).to_period('M').to_timestamp('M')
    return offset_date.strftime("%d-%b-%y")  # DD-MMM-YY format per Section 6.2


def _get_month_end_timestamp(date_str):
    """Convert a date-like string to a month-end Timestamp."""
    ts = pd.to_datetime(date_str, errors='coerce')
    if pd.isna(ts):
        return pd.NaT
    return ts.to_period('M').to_timestamp('M')


def _months_difference(start_date, end_date):
    """Whole-month difference between two Timestamps."""
    if pd.isna(start_date) or pd.isna(end_date):
        return 0
    return (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)


def _build_projected_inflation_lookup(comprehensive_results: Dict, valuation_date: str, deal_name: str) -> Tuple[Dict[str, float], float]:
    """
    DEPRECATED: This function builds inflation lookup at deal level, which is incorrect.
    Use _get_tranche_inflation_lookup() instead for tranche-specific inflation factors.
    
    This function is kept for reference but should not be used.
    It incorrectly picks an arbitrary tranche's inflation factors for the entire deal.
    
    Returns:
        Tuple of (lookup_dict, valuation_date_factor)
    """
    import warnings
    warnings.warn(
        "_build_projected_inflation_lookup is deprecated. Use _get_tranche_inflation_lookup instead.",
        DeprecationWarning,
        stacklevel=2
    )
    from const import get_deal_name_from_rga_tab
    
    val_month_end = _get_month_end_timestamp(valuation_date)
    
    # Find any tranche for this deal to get inflation factors
    for tranche_name, df in comprehensive_results.items():
        if get_deal_name_from_rga_tab(tranche_name) == deal_name:
            if df is None or df.empty:
                continue
            
            df_local = df.copy()
            df_local['Date_Datetime'] = pd.to_datetime(df_local['Date'], errors='coerce')
            
            # Build lookup dict
            lookup = {}
            for _, row in df_local.iterrows():
                date_str = row['Date']
                infl = pd.to_numeric(row.get('Projected_Infl_Factors', 1.0), errors='coerce')
                if not pd.isna(infl):
                    lookup[date_str] = float(infl)
            
            # Get valuation date factor
            val_row = df_local[df_local['Date_Datetime'] == val_month_end]
            if not val_row.empty:
                valdate_infl = pd.to_numeric(val_row.iloc[0]['Projected_Infl_Factors'], errors='coerce')
                if pd.isna(valdate_infl) or valdate_infl == 0:
                    valdate_infl = 1.0
            else:
                valdate_infl = 1.0
            
            return lookup, valdate_infl
    
    return {}, 1.0


def _get_tranche_inflation_lookup(
    comprehensive_results: Dict,
    tranche_name: str,
    valuation_date: str
) -> Tuple[Dict[str, float], float]:
    """
    Get inflation factors specific to a tranche from client data.
    
    Each tranche has its own Projected_Infl_Factors in the comprehensive_results.
    This function extracts them for the specific tranche.
    
    Args:
        comprehensive_results: Dict of {tranche_name: DataFrame with Projected_Infl_Factors}
        tranche_name: Name of the tranche to get inflation factors for
        valuation_date: Valuation date string (YYYY-MM-DD)
    
    Returns:
        Tuple of (lookup_dict, valuation_date_inflation_factor)
        - lookup_dict: Dict of {date_str: inflation_factor}
        - valuation_date_inflation_factor: Inflation factor at valuation date
    """
    val_month_end = _get_month_end_timestamp(valuation_date)
    
    df = comprehensive_results.get(tranche_name)
    if df is None or df.empty:
        logger.warning(f"No data found for tranche {tranche_name}, using default inflation factors")
        return {}, 1.0
    
    df_local = df.copy()
    df_local['Date_Datetime'] = pd.to_datetime(df_local['Date'], errors='coerce')
    
    # Build lookup dict from tranche's Projected_Infl_Factors
    lookup = {}
    for _, row in df_local.iterrows():
        date_str = row['Date']
        infl = pd.to_numeric(row.get('Projected_Infl_Factors', 1.0), errors='coerce')
        if not pd.isna(infl):
            lookup[date_str] = float(infl)
    
    # Get valuation date inflation factor
    val_row = df_local[df_local['Date_Datetime'] == val_month_end]
    if not val_row.empty:
        valdate_infl = pd.to_numeric(val_row.iloc[0].get('Projected_Infl_Factors', 1.0), errors='coerce')
        if pd.isna(valdate_infl) or valdate_infl == 0:
            valdate_infl = 1.0
    else:
        valdate_infl = 1.0
    
    logger.debug(f"Tranche {tranche_name}: Found {len(lookup)} inflation factors, valdate_infl={valdate_infl}")
    return lookup, valdate_infl


def _get_value_for_month_offset(df: pd.DataFrame, base_month_end, months_offset: int, column: str, fallback_to_first: bool = False):
    """
    Get a single column value for the row at base_month_end + months_offset months.
    
    Args:
        df: DataFrame with Date_Datetime column
        base_month_end: Base month-end timestamp
        months_offset: Number of months offset (can be negative)
        column: Column name to retrieve
        fallback_to_first: If True, fall back to first non-null value if target not found
    
    Returns:
        Value at target date or None
    """
    if df.empty:
        return None

    df_local = df.copy()
    if 'Date_Datetime' not in df_local.columns:
        df_local['Date_Datetime'] = pd.to_datetime(df_local['Date'], errors='coerce')
    df_local = df_local.sort_values('Date_Datetime')

    target_month = (base_month_end.to_period('M') + months_offset).to_timestamp('M')
    target_row = df_local.loc[df_local['Date_Datetime'] == target_month]

    if not target_row.empty:
        return pd.to_numeric(target_row.iloc[0][column], errors='coerce')

    if fallback_to_first:
        non_null = df_local.dropna(subset=[column])
        if not non_null.empty:
            return pd.to_numeric(non_null.iloc[0][column], errors='coerce')

    return None


def _safe_ratio(numerator, denominator):
    """Safe division that returns None if denominator is zero or any side is NaN."""
    num = pd.to_numeric(pd.Series([numerator]), errors='coerce').iloc[0]
    den = pd.to_numeric(pd.Series([denominator]), errors='coerce').iloc[0]
    if pd.isna(num) or pd.isna(den) or den == 0:
        return None
    return float(num) / float(den)


# =============================================================================
# Extended Claims Calculation Functions (Standard Method for Sensitivity)
# =============================================================================

def extend_total_actual_claims(
    tranche_df: pd.DataFrame,
    be_claims: pd.Series,
    stress_claims: pd.Series,
    valuation_date: str,
    grading_period: int,
    tranche_multiplier: float
) -> pd.DataFrame:
    """
    Extend Total_Actual_Claims for a tranche using Prophet projections for future dates.
    
    For dates <= valuation_date: Use existing Total_Actual_Claims from client data
    For dates > valuation_date: Project using scenario-specific blended Prophet claims
    
    Args:
        tranche_df: DataFrame with existing tranche data including Total_Actual_Claims
        be_claims: BE Prophet claims series (already adjusted by inflation and RGA share)
        stress_claims: 1in200 Prophet claims series (already adjusted by inflation and RGA share)
        valuation_date: Valuation date string (YYYY-MM-DD)
        grading_period: Grading period for blending (0 for Scenario 1, 60 for others)
        tranche_multiplier: Multiplier to allocate deal claims to this tranche
    
    Returns:
        DataFrame with extended Total_Actual_Claims for all dates
    """
    if tranche_df is None or tranche_df.empty:
        return pd.DataFrame()
    
    val_month_end = _get_month_end_timestamp(valuation_date)
    
    # Make a copy and ensure Date_Datetime column exists
    df = tranche_df.copy()
    df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('Date_Datetime').reset_index(drop=True)
    
    # Get all dates from the tranche data
    all_dates = df['Date'].tolist()
    all_date_dts = df['Date_Datetime'].tolist()
    
    # For each date, determine if we need to project Total_Actual_Claims
    extended_claims = []
    
    for i, (date_str, date_dt) in enumerate(zip(all_dates, all_date_dts)):
        if pd.isna(date_dt):
            extended_claims.append(0.0)
            continue
        
        # For dates <= valuation date, use existing Total_Actual_Claims
        if date_dt <= val_month_end:
            existing_claim = pd.to_numeric(df.iloc[i].get('Total_Actual_Claims', 0.0), errors='coerce')
            extended_claims.append(float(existing_claim) if not pd.isna(existing_claim) else 0.0)
        else:
            # For future dates, project using Prophet claims
            # Calculate offset from valuation date
            offset = _months_difference(val_month_end, date_dt)
            
            # Calculate blend factor based on offset and grading period
            if grading_period == 0:
                blend_factor = 0.0  # 100% BE
            else:
                blend_factor = min(1.0, offset / grading_period)
            
            # Get BE and Stress claims for this date
            be_val = be_claims.get(date_str, 0.0) if date_str in be_claims.index else 0.0
            stress_val = stress_claims.get(date_str, 0.0) if date_str in stress_claims.index else 0.0
            
            # Blend claims
            blended_claim = (1 - blend_factor) * be_val + blend_factor * stress_val
            
            # Apply tranche multiplier to get tranche-level claims
            tranche_claim = blended_claim * tranche_multiplier
            
            extended_claims.append(tranche_claim)
    
    # Update the Total_Actual_Claims column
    df['Total_Actual_Claims'] = extended_claims
    
    return df


def precompute_blending_vectors(
    tranche_df: pd.DataFrame,
    be_claims: pd.Series,
    stress_claims: pd.Series,
    valuation_date: str,
    tranche_multiplier: float,
    tranche_infl_lookup: Dict[str, float],
    tranche_valdate_infl: float,
    adjust_to_actual_claim: float = 1.0
) -> Tuple[pd.DataFrame, np.ndarray, np.ndarray]:
    """
    Pre-compute base and delta vectors for efficient blending across offsets.
    
    For dates <= valuation_date: Use existing Total_Actual_Claims (stored in base, delta=0)
    For dates > valuation_date: 
        base = (BE / adjust_to_actual_claim) * multiplier * infl_ratio
        delta = ((1in200 - BE) / adjust_to_actual_claim) * multiplier * infl_ratio
    
    Where infl_ratio = projected_infl_factor / valdate_infl_factor (per tranche)
    
    This allows efficient calculation of blended claims for any offset:
        Total_Actual_Claims = base + blend_factor * delta
    
    Where blend_factor = min(1.0, offset / grading_period)
    
    Args:
        tranche_df: DataFrame with existing tranche data including Total_Actual_Claims
        be_claims: BE Prophet claims series (adjusted by RGA share only, NOT inflation)
        stress_claims: 1in200 Prophet claims series (adjusted by RGA share only, NOT inflation)
        valuation_date: Valuation date string (YYYY-MM-DD)
        tranche_multiplier: Multiplier to allocate deal claims to this tranche
        tranche_infl_lookup: Dict of {date_str: inflation_factor} for this specific tranche
        tranche_valdate_infl: Valuation date inflation factor for this specific tranche
        adjust_to_actual_claim: Scalar adjustment factor computed per tranche to calibrate
            prophet claims to actual claims at valuation date (default 1.0 = no adjustment)
    
    Returns:
        Tuple of (prepared_df, base_vector, delta_vector)
    """
    if tranche_df is None or tranche_df.empty:
        return pd.DataFrame(), np.array([]), np.array([])
    
    val_month_end = _get_month_end_timestamp(valuation_date)
    
    # Make a copy and ensure Date_Datetime column exists
    df = tranche_df.copy()
    df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('Date_Datetime').reset_index(drop=True)
    
    # Get all dates from the tranche data
    all_dates = df['Date'].tolist()
    all_date_dts = df['Date_Datetime'].tolist()
    
    # Pre-compute base and delta vectors
    base_vector = []
    delta_vector = []
    
    for i, (date_str, date_dt) in enumerate(zip(all_dates, all_date_dts)):
        if pd.isna(date_dt):
            base_vector.append(0.0)
            delta_vector.append(0.0)
            continue
        
        # For dates <= valuation date, use existing Total_Actual_Claims
        if date_dt <= val_month_end:
            existing_claim = pd.to_numeric(df.iloc[i].get('Total_Actual_Claims', 0.0), errors='coerce')
            base_vector.append(float(existing_claim) if not pd.isna(existing_claim) else 0.0)
            delta_vector.append(0.0)  # No blending for historical dates
        else:
            # For future dates, compute base (BE) and delta (1in200 - BE)
            # Apply tranche-specific inflation adjustment
            be_val = be_claims.get(date_str, 0.0) if date_str in be_claims.index else 0.0
            stress_val = stress_claims.get(date_str, 0.0) if date_str in stress_claims.index else 0.0
            
            # Get inflation ratio for this date (tranche-specific)
            infl_factor = tranche_infl_lookup.get(date_str, tranche_valdate_infl)
            if pd.isna(infl_factor) or infl_factor == 0:
                infl_factor = tranche_valdate_infl
            infl_ratio = infl_factor / tranche_valdate_infl if tranche_valdate_infl != 0 else 1.0
            
            # Order: prophet / rga_share (already done) / adjust_to_actual_claim * multiplier * infl_ratio
            base_vector.append((be_val / adjust_to_actual_claim) * tranche_multiplier * infl_ratio)
            delta_vector.append(((stress_val - be_val) / adjust_to_actual_claim) * tranche_multiplier * infl_ratio)
    
    return df, np.array(base_vector), np.array(delta_vector)


def apply_blend_factor(
    prepared_df: pd.DataFrame,
    base_vector: np.ndarray,
    delta_vector: np.ndarray,
    blend_factor: float
) -> pd.DataFrame:
    """
    Apply blend factor to get Total_Actual_Claims for a specific offset.
    
    Formula: Total_Actual_Claims = base + blend_factor * delta
    
    This is equivalent to: (1 - blend_factor) * BE + blend_factor * 1in200
    
    Args:
        prepared_df: DataFrame prepared by precompute_blending_vectors
        base_vector: Pre-computed base values (BE * multiplier for future dates)
        delta_vector: Pre-computed delta values ((1in200 - BE) * multiplier for future dates)
        blend_factor: Blend factor based on sensitivity offset (offset / grading_period)
    
    Returns:
        DataFrame with Total_Actual_Claims updated for this blend factor
    """
    if prepared_df is None or prepared_df.empty:
        return pd.DataFrame()
    
    df = prepared_df.copy()
    
    # Vectorized calculation: Total_Actual_Claims = base + blend_factor * delta
    df['Total_Actual_Claims'] = base_vector + blend_factor * delta_vector
    
    return df


def compute_factors_for_offset(
    extended_df: pd.DataFrame,
    anchor_date: str,
    inception_date: pd.Timestamp
) -> Dict:
    """
    Compute Experience Factor, Adjustment Factor, and related factors for a given anchor date.
    
    This mirrors the standard calculation in calculation.py but parameterized by anchor date
    instead of valuation date.
    
    Args:
        extended_df: DataFrame with extended Total_Actual_Claims and vector columns
        anchor_date: The anchor date (valuation_date + offset) as string
        inception_date: Inception date of the tranche
    
    Returns:
        Dict containing:
            - Experience_Factor
            - Interpolation_Vector
            - Adjustment_Factor
            - Credibility_Factor
            - X_Actual, X_Fixed, X_Increased, X_Decreased
    """
    if extended_df.empty:
        return {
            'Experience_Factor': 0.0,
            'Interpolation_Vector': 0,
            'Adjustment_Factor': 1.0,
            'Credibility_Factor': 0.0,
            'X_Actual': None,
            'X_Fixed': None,
            'X_Increased': None,
            'X_Decreased': None,
        }
    
    df = extended_df.copy()
    if 'Date_Datetime' not in df.columns:
        df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('Date_Datetime')
    
    anchor_month_end = _get_month_end_timestamp(anchor_date)
    
    # Calculate credibility factor based on months since inception to anchor date
    months_since_inception = _months_difference(inception_date, anchor_month_end) + 1
    raw_months = max(months_since_inception - 6, 0)
    capped_months = min(raw_months, 60)
    credibility_factor = (capped_months / 60.0) ** 2 if capped_months > 0 else 0.0
    
    # Get values at anchor_date - 66 and anchor_date - 6 for X calculations
    c66_actual = _get_value_for_month_offset(df, anchor_month_end, -66, 'Total_Actual_Claims', fallback_to_first=True)
    c6_actual = _get_value_for_month_offset(df, anchor_month_end, -6, 'Total_Actual_Claims', fallback_to_first=False)
    c66_fixed = _get_value_for_month_offset(df, anchor_month_end, -66, 'Fixed_w_Real', fallback_to_first=True)
    c6_fixed = _get_value_for_month_offset(df, anchor_month_end, -6, 'Fixed_w_Real', fallback_to_first=False)
    c66_increased = _get_value_for_month_offset(df, anchor_month_end, -66, 'Increased_w_Real', fallback_to_first=True)
    c6_increased = _get_value_for_month_offset(df, anchor_month_end, -6, 'Increased_w_Real', fallback_to_first=False)
    c66_decreased = _get_value_for_month_offset(df, anchor_month_end, -66, 'Decreased_w_Real', fallback_to_first=True)
    c6_decreased = _get_value_for_month_offset(df, anchor_month_end, -6, 'Decreased_w_Real', fallback_to_first=False)
    
    # Calculate X ratios
    numerator_actual = c66_actual - c6_actual if c66_actual is not None and c6_actual is not None else None
    x_actual = _safe_ratio(numerator_actual, c66_actual)
    numerator_fixed = c66_fixed - c6_fixed if c66_fixed is not None and c6_fixed is not None else None
    x_fixed = _safe_ratio(numerator_fixed, c66_fixed)
    numerator_increased = c66_increased - c6_increased if c66_increased is not None and c6_increased is not None else None
    x_increased = _safe_ratio(numerator_increased, c66_increased)
    numerator_decreased = c66_decreased - c6_decreased if c66_decreased is not None and c6_decreased is not None else None
    x_decreased = _safe_ratio(numerator_decreased, c66_decreased)
    
    # Calculate experience_raw
    experience_raw = 0.0
    if all(v is not None for v in [x_actual, x_fixed, x_increased, x_decreased]):
        if abs(x_actual - x_fixed) < 1e-6:
            experience_raw = 0.0
        elif x_fixed < x_actual:
            denom = (x_fixed - x_decreased)
            experience_raw = 0.0 if abs(denom) < 1e-6 else (x_fixed - x_actual) / denom
        else:
            denom = (x_increased - x_fixed)
            experience_raw = 0.0 if abs(denom) < 1e-6 else (x_actual - x_fixed) / denom
    
    experience_factor = experience_raw * credibility_factor
    
    # Determine interpolation vector
    if x_fixed is None or x_actual is None:
        interpolation_vector = 0
    else:
        interpolation_vector = 0 if x_fixed < x_actual else 1
    
    # Calculate adjustment factor using sums at anchor_date - 8, -7, -6
    offsets = [-8, -7, -6]
    sum_claims = 0.0
    sum_fixed_real = 0.0
    sum_increased_real = 0.0
    sum_decreased_real = 0.0
    
    for off in offsets:
        month_end = (anchor_month_end.to_period('M') + off).to_timestamp('M')
        row = df.loc[df['Date_Datetime'] == month_end]
        if row.empty:
            continue
        r = row.iloc[0]
        sum_claims += pd.to_numeric(r.get('Total_Actual_Claims', 0.0), errors='coerce') or 0.0
        sum_fixed_real += pd.to_numeric(r.get('Fixed_w_Real', 0.0), errors='coerce') or 0.0
        sum_increased_real += pd.to_numeric(r.get('Increased_w_Real', 0.0), errors='coerce') or 0.0
        sum_decreased_real += pd.to_numeric(r.get('Decreased_w_Real', 0.0), errors='coerce') or 0.0
    
    if sum_claims == 0 and sum_fixed_real == 0 and sum_increased_real == 0 and sum_decreased_real == 0:
        adjustment_factor = 1.0
    else:
        base_denominator = (1.0 - experience_factor) * sum_fixed_real
        if interpolation_vector == 0:
            base_denominator += experience_factor * sum_decreased_real
        else:
            base_denominator += experience_factor * sum_increased_real
        
        adjustment_factor = sum_claims / base_denominator if base_denominator != 0 else 1.0
    
    return {
        'Experience_Factor': experience_factor,
        'Interpolation_Vector': interpolation_vector,
        'Adjustment_Factor': adjustment_factor,
        'Credibility_Factor': credibility_factor,
        'X_Actual': x_actual,
        'X_Fixed': x_fixed,
        'X_Increased': x_increased,
        'X_Decreased': x_decreased,
    }


def compute_adjusted_float_vector_for_offset(
    extended_df: pd.DataFrame,
    factors: Dict,
    anchor_date: str
) -> np.ndarray:
    """
    Compute Adjusted_Float_Vector using the standard formula for dates >= anchor_date.
    
    Formula:
        mix_vector = Decreased_w_Original if interpolation_vector == 0 else Increased_w_Original
        Float_Vector = ((1 - experience_factor) * Fixed_w_Original + experience_factor * mix_vector) * adjustment_factor
        Adjusted_Float_Vector = Float_Vector * (Projected_Infl_Factors / Historical_Infl_Factors)
    
    Args:
        extended_df: DataFrame with extended data including vector columns
        factors: Dict from compute_factors_for_offset()
        anchor_date: The anchor date (valuation_date + offset) as string
    
    Returns:
        NumPy array of Adjusted_Float_Vector values for dates >= anchor_date
    """
    if extended_df.empty:
        return np.array([])
    
    df = extended_df.copy()
    if 'Date_Datetime' not in df.columns:
        df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('Date_Datetime')
    
    anchor_month_end = _get_month_end_timestamp(anchor_date)
    
    # Filter to dates >= anchor date
    df_filtered = df[df['Date_Datetime'] >= anchor_month_end].copy()
    
    if df_filtered.empty:
        return np.array([])
    
    # Extract factors
    experience_factor = factors['Experience_Factor']
    interpolation_vector = factors['Interpolation_Vector']
    adjustment_factor = factors['Adjustment_Factor']
    
    # Get vector columns
    fixed_orig = pd.to_numeric(df_filtered.get('Fixed_w_Original', pd.Series()), errors='coerce').fillna(0).values
    increased_orig = pd.to_numeric(df_filtered.get('Increased_w_Original', pd.Series()), errors='coerce').fillna(0).values
    decreased_orig = pd.to_numeric(df_filtered.get('Decreased_w_Original', pd.Series()), errors='coerce').fillna(0).values
    
    hist_factors = pd.to_numeric(df_filtered.get('Historical_Infl_Factors', pd.Series()), errors='coerce').fillna(1).values
    proj_factors = pd.to_numeric(df_filtered.get('Projected_Infl_Factors', pd.Series()), errors='coerce').fillna(1).values
    
    # Determine mix vector based on interpolation
    mix_vector = decreased_orig if interpolation_vector == 0 else increased_orig
    
    # Calculate Float_Vector
    float_vector = ((1.0 - experience_factor) * fixed_orig + experience_factor * mix_vector) * adjustment_factor
    
    # Calculate inflation ratio (handle division by zero)
    with np.errstate(divide='ignore', invalid='ignore'):
        inflation_ratio = np.where(hist_factors != 0, proj_factors / hist_factors, 1.0)
        inflation_ratio = np.nan_to_num(inflation_ratio, nan=1.0, posinf=1.0, neginf=1.0)
    
    # Calculate Adjusted_Float_Vector
    adjusted_float_vector = float_vector * inflation_ratio
    
    return adjusted_float_vector


# =============================================================================
# Optimized Helper Functions for Sensitivity Analysis
# =============================================================================

def _prepare_tranche_arrays(
    tranche_df: pd.DataFrame,
    valuation_date: str
) -> Dict:
    """
    Pre-compute all arrays needed for tranche calculations.
    
    Converts DataFrame columns to NumPy arrays once, avoiding repeated
    pd.to_numeric() calls in hot loops.
    
    Args:
        tranche_df: DataFrame with tranche data
        valuation_date: Valuation date string
    
    Returns:
        Dict with pre-computed arrays and metadata
    """
    if tranche_df is None or tranche_df.empty:
        return None
    
    df = tranche_df.copy()
    df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('Date_Datetime').reset_index(drop=True)
    
    val_month_end = _get_month_end_timestamp(valuation_date)
    
    # Pre-compute all arrays
    dates = df['Date'].values
    date_datetimes = df['Date_Datetime'].values
    
    # Find valuation date index for fast slicing
    val_idx = 0
    for i, dt in enumerate(date_datetimes):
        if pd.Timestamp(dt) >= val_month_end:
            val_idx = i
            break
    
    # Pre-convert all numeric columns to arrays
    fixed_w_real = pd.to_numeric(df.get('Fixed_w_Real', pd.Series()), errors='coerce').fillna(0).values
    fixed_w_original = pd.to_numeric(df.get('Fixed_w_Original', pd.Series()), errors='coerce').fillna(0).values
    increased_w_original = pd.to_numeric(df.get('Increased_w_Original', pd.Series()), errors='coerce').fillna(0).values
    decreased_w_original = pd.to_numeric(df.get('Decreased_w_Original', pd.Series()), errors='coerce').fillna(0).values
    increased_w_real = pd.to_numeric(df.get('Increased_w_Real', pd.Series()), errors='coerce').fillna(0).values
    decreased_w_real = pd.to_numeric(df.get('Decreased_w_Real', pd.Series()), errors='coerce').fillna(0).values
    deferred_w_real = pd.to_numeric(df.get('Deferred_w_Real', pd.Series()), errors='coerce').fillna(0).values
    pensioner_w_real = pd.to_numeric(df.get('Pensioner_w_Real', pd.Series()), errors='coerce').fillna(0).values
    total_actual_claims = pd.to_numeric(df.get('Total_Actual_Claims', pd.Series()), errors='coerce').fillna(0).values
    hist_infl_factors = pd.to_numeric(df.get('Historical_Infl_Factors', pd.Series()), errors='coerce').fillna(1).values
    proj_infl_factors = pd.to_numeric(df.get('Projected_Infl_Factors', pd.Series()), errors='coerce').fillna(1).values
    
    # Pre-compute inflation ratio
    with np.errstate(divide='ignore', invalid='ignore'):
        inflation_ratio = np.where(hist_infl_factors != 0, proj_infl_factors / hist_infl_factors, 1.0)
        inflation_ratio = np.nan_to_num(inflation_ratio, nan=1.0, posinf=1.0, neginf=1.0)
    
    # Build date to index mapping for fast lookups
    date_to_idx = {}
    for i, dt in enumerate(date_datetimes):
        if pd.notna(dt):
            date_to_idx[pd.Timestamp(dt)] = i
    
    return {
        'df': df,
        'dates': dates,
        'date_datetimes': date_datetimes,
        'date_to_idx': date_to_idx,
        'val_idx': val_idx,
        'val_month_end': val_month_end,
        'fixed_w_real': fixed_w_real,
        'fixed_w_original': fixed_w_original,
        'increased_w_original': increased_w_original,
        'decreased_w_original': decreased_w_original,
        'increased_w_real': increased_w_real,
        'decreased_w_real': decreased_w_real,
        'deferred_w_real': deferred_w_real,
        'pensioner_w_real': pensioner_w_real,
        'total_actual_claims': total_actual_claims,
        'hist_infl_factors': hist_infl_factors,
        'proj_infl_factors': proj_infl_factors,
        'inflation_ratio': inflation_ratio,
        'n_rows': len(df),
    }


def _prepare_fee_vector_arrays(
    fee_vectors_data: Dict,
    valuation_date: str
) -> Dict[str, List[Dict]]:
    """
    Pre-compute NumPy arrays and date-to-index mappings for additional fee vectors.

    Converts the raw fee vector DataFrames (Am/In per deal) into a format
    suitable for fast lookups in the sensitivity hot loop.

    Args:
        fee_vectors_data: {deal_name: {'Am': DataFrame, 'In': DataFrame}}
            where each DataFrame has 'Date' and 'Fee_Vector' columns.
        valuation_date: Valuation date string (YYYY-MM-DD)

    Returns:
        {deal_name: [{'fee_vector': np.ndarray, 'date_to_idx': dict, 'date_datetimes': np.ndarray}, ...]}
    """
    result: Dict[str, List[Dict]] = {}

    for deal_name, fee_data_by_type in fee_vectors_data.items():
        deal_arrays = []
        for fee_type, fee_df in fee_data_by_type.items():
            if fee_df is None or fee_df.empty:
                continue

            df = fee_df.copy()
            df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
            df = df.sort_values('Date_Datetime').reset_index(drop=True)

            fee_values = pd.to_numeric(df['Fee_Vector'], errors='coerce').fillna(0).values
            date_datetimes = df['Date_Datetime'].values

            date_to_idx = {}
            for i, dt in enumerate(date_datetimes):
                if pd.notna(dt):
                    date_to_idx[pd.Timestamp(dt)] = i

            deal_arrays.append({
                'fee_vector': fee_values,
                'date_to_idx': date_to_idx,
                'date_datetimes': date_datetimes,
                'n_rows': len(df),
                'fee_type': fee_type,
            })

        if deal_arrays:
            result[deal_name] = deal_arrays
            logger.debug(f"Pre-processed {len(deal_arrays)} fee vector arrays for {deal_name}")

    return result


def _compute_factors_vectorized(
    tranche_arrays: Dict,
    anchor_month_end: pd.Timestamp,
    inception_date: pd.Timestamp,
    blended_claims: np.ndarray
) -> Dict:
    """
    Compute Experience Factor, Adjustment Factor using pre-computed arrays.
    
    This is a vectorized version of compute_factors_for_offset() that works
    with pre-computed arrays instead of DataFrames.
    
    Args:
        tranche_arrays: Pre-computed arrays from _prepare_tranche_arrays()
        anchor_month_end: The anchor date timestamp
        inception_date: Inception date of the tranche
        blended_claims: Total_Actual_Claims array with blended values for future dates
    
    Returns:
        Dict with computed factors
    """
    date_to_idx = tranche_arrays['date_to_idx']
    fixed_w_real = tranche_arrays['fixed_w_real']
    increased_w_real = tranche_arrays['increased_w_real']
    decreased_w_real = tranche_arrays['decreased_w_real']
    
    # Calculate credibility factor
    months_since_inception = _months_difference(inception_date, anchor_month_end) + 1
    raw_months = max(months_since_inception - 6, 0)
    capped_months = min(raw_months, 60)
    credibility_factor = (capped_months / 60.0) ** 2 if capped_months > 0 else 0.0
    
    # Helper to get value at offset
    def get_val_at_offset(array, offset):
        target = (anchor_month_end.to_period('M') + offset).to_timestamp('M')
        idx = date_to_idx.get(target)
        if idx is not None:
            return array[idx]
        return None
    
    # Get values for X calculations at -66 and -6
    c66_actual = get_val_at_offset(blended_claims, -66)
    c6_actual = get_val_at_offset(blended_claims, -6)
    c66_fixed = get_val_at_offset(fixed_w_real, -66)
    c6_fixed = get_val_at_offset(fixed_w_real, -6)
    c66_increased = get_val_at_offset(increased_w_real, -66)
    c6_increased = get_val_at_offset(increased_w_real, -6)
    c66_decreased = get_val_at_offset(decreased_w_real, -66)
    c6_decreased = get_val_at_offset(decreased_w_real, -6)
    
    # Fallback for -66 values if not found
    if c66_actual is None:
        c66_actual = blended_claims[0] if len(blended_claims) > 0 else 0.0
    if c66_fixed is None:
        c66_fixed = fixed_w_real[0] if len(fixed_w_real) > 0 else 0.0
    if c66_increased is None:
        c66_increased = increased_w_real[0] if len(increased_w_real) > 0 else 0.0
    if c66_decreased is None:
        c66_decreased = decreased_w_real[0] if len(decreased_w_real) > 0 else 0.0
    
    # Calculate X ratios
    def safe_ratio(num, den):
        if num is None or den is None or den == 0 or np.isnan(num) or np.isnan(den):
            return None
        return float(num) / float(den)
    
    numerator_actual = c66_actual - c6_actual if c66_actual is not None and c6_actual is not None else None
    x_actual = safe_ratio(numerator_actual, c66_actual)
    numerator_fixed = c66_fixed - c6_fixed if c66_fixed is not None and c6_fixed is not None else None
    x_fixed = safe_ratio(numerator_fixed, c66_fixed)
    numerator_increased = c66_increased - c6_increased if c66_increased is not None and c6_increased is not None else None
    x_increased = safe_ratio(numerator_increased, c66_increased)
    numerator_decreased = c66_decreased - c6_decreased if c66_decreased is not None and c6_decreased is not None else None
    x_decreased = safe_ratio(numerator_decreased, c66_decreased)
    
    # Calculate experience_raw
    experience_raw = 0.0
    if all(v is not None for v in [x_actual, x_fixed, x_increased, x_decreased]):
        if abs(x_actual - x_fixed) < 1e-6:
            experience_raw = 0.0
        elif x_fixed < x_actual:
            denom = (x_fixed - x_decreased)
            experience_raw = 0.0 if abs(denom) < 1e-6 else (x_fixed - x_actual) / denom
        else:
            denom = (x_increased - x_fixed)
            experience_raw = 0.0 if abs(denom) < 1e-6 else (x_actual - x_fixed) / denom
    
    experience_factor = experience_raw * credibility_factor
    
    # Interpolation vector
    if x_fixed is None or x_actual is None:
        interpolation_vector = 0
    else:
        interpolation_vector = 0 if x_fixed < x_actual else 1
    
    # Calculate adjustment factor from sums at -8, -7, -6
    sum_claims = 0.0
    sum_fixed_real = 0.0
    sum_increased_real = 0.0
    sum_decreased_real = 0.0
    
    for off in [-8, -7, -6]:
        target = (anchor_month_end.to_period('M') + off).to_timestamp('M')
        idx = date_to_idx.get(target)
        if idx is not None:
            sum_claims += blended_claims[idx] if not np.isnan(blended_claims[idx]) else 0.0
            sum_fixed_real += fixed_w_real[idx] if not np.isnan(fixed_w_real[idx]) else 0.0
            sum_increased_real += increased_w_real[idx] if not np.isnan(increased_w_real[idx]) else 0.0
            sum_decreased_real += decreased_w_real[idx] if not np.isnan(decreased_w_real[idx]) else 0.0
    
    if sum_claims == 0 and sum_fixed_real == 0 and sum_increased_real == 0 and sum_decreased_real == 0:
        adjustment_factor = 1.0
    else:
        base_denominator = (1.0 - experience_factor) * sum_fixed_real
        if interpolation_vector == 0:
            base_denominator += experience_factor * sum_decreased_real
        else:
            base_denominator += experience_factor * sum_increased_real
        adjustment_factor = sum_claims / base_denominator if base_denominator != 0 else 1.0
    
    return {
        'Experience_Factor': experience_factor,
        'Interpolation_Vector': interpolation_vector,
        'Adjustment_Factor': adjustment_factor,
        'Credibility_Factor': credibility_factor,
    }


def _compute_adjusted_float_vector_fast(
    tranche_arrays: Dict,
    factors: Dict,
    start_idx: int
) -> np.ndarray:
    """
    Compute Adjusted_Float_Vector using pre-computed arrays.
    
    Args:
        tranche_arrays: Pre-computed arrays from _prepare_tranche_arrays()
        factors: Factors dict from _compute_factors_vectorized()
        start_idx: Index to start from (corresponding to anchor date)
    
    Returns:
        NumPy array of Adjusted_Float_Vector values
    """
    fixed_orig = tranche_arrays['fixed_w_original'][start_idx:]
    increased_orig = tranche_arrays['increased_w_original'][start_idx:]
    decreased_orig = tranche_arrays['decreased_w_original'][start_idx:]
    inflation_ratio = tranche_arrays['inflation_ratio'][start_idx:]
    
    experience_factor = factors['Experience_Factor']
    interpolation_vector = factors['Interpolation_Vector']
    adjustment_factor = factors['Adjustment_Factor']
    
    # Mix vector based on interpolation
    mix_vector = decreased_orig if interpolation_vector == 0 else increased_orig
    
    # Float_Vector = ((1 - exp_factor) * fixed + exp_factor * mix) * adj_factor
    float_vector = ((1.0 - experience_factor) * fixed_orig + experience_factor * mix_vector) * adjustment_factor
    
    # Adjusted_Float_Vector = Float_Vector * inflation_ratio
    return float_vector * inflation_ratio


def _precompute_blending_arrays(
    tranche_arrays: Dict,
    be_claims: pd.Series,
    stress_claims: pd.Series,
    tranche_multiplier: float,
    tranche_infl_lookup: Dict[str, float],
    tranche_valdate_infl: float,
    adjust_to_actual_claim: float = 1.0
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Pre-compute base and delta arrays for efficient blending.
    
    Returns arrays that can be used to compute blended claims for any blend factor:
        blended_claims = base + blend_factor * delta
    
    Mathematical Equivalence:
    -------------------------
    This implementation is mathematically equivalent to the specified order:
    
    Specified order:
        1. Blend: blended_prophet = (1 - bf) * BE + bf * 1in200
        2. Apply fixed multiplier: tranche_claims = blended / rga_share * multiplier
        3. Adjust to actual claim: adjusted = tranche_claims / adjust_to_actual_claim
        4. Inflate per tranche: final = adjusted * (proj_infl / valdate_infl)
    
    This implementation (where BE/1in200 are already divided by rga_share):
        base = (BE/rga / adj) * mult * infl_ratio
        delta = ((1in200 - BE)/rga / adj) * mult * infl_ratio
        blended = base + bf * delta
    
    Args:
        tranche_arrays: Pre-computed arrays from _prepare_tranche_arrays()
        be_claims: BE Prophet claims series (already divided by rga_share at deal level)
        stress_claims: 1in200 Prophet claims series (already divided by rga_share at deal level)
        tranche_multiplier: Multiplier for this tranche (fixed, from prior month claims)
        tranche_infl_lookup: Inflation factor lookup (tranche-specific from client data)
        tranche_valdate_infl: Valuation date inflation factor (tranche-specific)
        adjust_to_actual_claim: Scalar adjustment factor computed per tranche to calibrate
            prophet claims to actual claims at valuation date (default 1.0 = no adjustment)
    
    Returns:
        Tuple of (base_array, delta_array)
    """
    dates = tranche_arrays['dates']
    date_datetimes = tranche_arrays['date_datetimes']
    val_month_end = tranche_arrays['val_month_end']
    total_actual_claims = tranche_arrays['total_actual_claims']
    n_rows = tranche_arrays['n_rows']
    
    base_array = np.zeros(n_rows)
    delta_array = np.zeros(n_rows)
    
    for i in range(n_rows):
        date_str = dates[i]
        date_dt = date_datetimes[i]
        
        if pd.isna(date_dt):
            continue
        
        if pd.Timestamp(date_dt) <= val_month_end:
            # Historical: use actual claims
            base_array[i] = total_actual_claims[i]
            delta_array[i] = 0.0
        else:
            # Future: compute from Prophet claims
            be_val = be_claims.get(date_str, 0.0) if date_str in be_claims.index else 0.0
            stress_val = stress_claims.get(date_str, 0.0) if date_str in stress_claims.index else 0.0
            
            # Get inflation ratio
            infl_factor = tranche_infl_lookup.get(date_str, tranche_valdate_infl)
            if pd.isna(infl_factor) or infl_factor == 0:
                infl_factor = tranche_valdate_infl
            infl_ratio = infl_factor / tranche_valdate_infl if tranche_valdate_infl != 0 else 1.0
            
            # Order: prophet / rga_share (already done) / adjust_to_actual_claim * multiplier * infl_ratio
            base_array[i] = (be_val / adjust_to_actual_claim) * tranche_multiplier * infl_ratio
            delta_array[i] = ((stress_val - be_val) / adjust_to_actual_claim) * tranche_multiplier * infl_ratio
    
    return base_array, delta_array


def _get_anchor_date_index(
    tranche_arrays: Dict,
    offset: int
) -> int:
    """
    Get the array index corresponding to anchor date (valuation + offset months).
    
    Args:
        tranche_arrays: Pre-computed arrays
        offset: Offset in months from valuation date
    
    Returns:
        Array index, or -1 if not found
    """
    val_month_end = tranche_arrays['val_month_end']
    anchor = (val_month_end.to_period('M') + offset).to_timestamp('M')
    return tranche_arrays['date_to_idx'].get(anchor, -1)


def _compute_adjust_to_actual_claim(
    tranche_df: pd.DataFrame,
    be_claims_adj: pd.Series,
    tranche_multiplier: float,
    valuation_date: str
) -> float:
    """
    Compute the adjust_to_actual_claim factor for a tranche.

    This calibrates prophet projections to match the observed actual claim
    at the valuation date. Computed once per tranche and applied to all
    future prophet claims (both BE and stress) before inflation adjustment.

    Formula:
        1st_proj_claims = Actual_Claims[V]
                          * Fixed_Vectors[V+1]^2
                          / Fixed_Vectors[V]
                          / Fixed_Vectors[V+2]
                          * Prophet_tranche[V+2]
                          / Prophet_tranche[V+1]

        adjust_to_actual_claim = Prophet_tranche[V+1] / 1st_proj_claims

    Where V = valuation date, Prophet_tranche = BE claims / rga_share * multiplier.

    Args:
        tranche_df: DataFrame with Total_Actual_Claims and Fixed_Vectors columns
        be_claims_adj: BE Prophet claims series (already divided by rga_share),
                       indexed by date string (YYYY-MM-DD)
        tranche_multiplier: Multiplier to allocate deal claims to this tranche
        valuation_date: Valuation date string (YYYY-MM-DD)

    Returns:
        Scalar adjustment factor. Falls back to 1.0 if any required data
        is missing or a denominator is zero.
    """
    if tranche_df is None or tranche_df.empty:
        logger.warning("Empty tranche DataFrame - adjust_to_actual_claim defaulting to 1.0")
        return 1.0

    val_month_end = _get_month_end_timestamp(valuation_date)
    if pd.isna(val_month_end):
        logger.warning(f"Invalid valuation date '{valuation_date}' - defaulting to 1.0")
        return 1.0

    df = tranche_df.copy()
    if 'Date_Datetime' not in df.columns:
        df['Date_Datetime'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.sort_values('Date_Datetime')

    # Helper: get a column value at a specific month offset from val_month_end
    def _get_val(column: str, month_offset: int):
        target = (val_month_end.to_period('M') + month_offset).to_timestamp('M')
        row = df.loc[df['Date_Datetime'] == target]
        if row.empty:
            return None
        val = pd.to_numeric(row.iloc[0].get(column, None), errors='coerce')
        return val if not pd.isna(val) else None

    # --- Gather required data points ---

    # Actual_Claims at ValDate (from client file, already divided by RGA share)
    actual_claims_valdate = _get_val('Total_Actual_Claims', 0)
    if actual_claims_valdate is None or actual_claims_valdate == 0:
        logger.warning("Actual_Claims at ValDate is missing or zero - defaulting to 1.0")
        return 1.0

    # Fixed_Vectors at ValDate, ValDate+1, ValDate+2 (raw, from consolidated RGA file)
    fv_v = _get_val('Fixed_Vectors', 0)
    fv_v1 = _get_val('Fixed_Vectors', 1)
    fv_v2 = _get_val('Fixed_Vectors', 2)

    if any(v is None or v == 0 for v in [fv_v, fv_v1, fv_v2]):
        logger.warning(
            f"Fixed_Vectors at ValDate/+1/+2 missing or zero "
            f"(fv_v={fv_v}, fv_v1={fv_v1}, fv_v2={fv_v2}) - defaulting to 1.0"
        )
        return 1.0

    # Prophet BE claims per tranche at ValDate+1 and ValDate+2
    vp1_date = (val_month_end.to_period('M') + 1).to_timestamp('M').strftime('%Y-%m-%d')
    vp2_date = (val_month_end.to_period('M') + 2).to_timestamp('M').strftime('%Y-%m-%d')

    prophet_be_vp1 = be_claims_adj.get(vp1_date, 0.0) if vp1_date in be_claims_adj.index else 0.0
    prophet_be_vp2 = be_claims_adj.get(vp2_date, 0.0) if vp2_date in be_claims_adj.index else 0.0

    prophet_tranche_vp1 = prophet_be_vp1 * tranche_multiplier
    prophet_tranche_vp2 = prophet_be_vp2 * tranche_multiplier

    if prophet_tranche_vp1 == 0 or prophet_tranche_vp2 == 0:
        logger.warning(
            f"Prophet BE tranche claims at ValDate+1/+2 are zero "
            f"(vp1={prophet_tranche_vp1}, vp2={prophet_tranche_vp2}) - defaulting to 1.0"
        )
        return 1.0

    # --- Compute 1st_proj_claims ---
    first_proj_claims = (
        actual_claims_valdate
        * (fv_v1 ** 2)
        / fv_v
        / fv_v2
        * prophet_tranche_vp2
        / prophet_tranche_vp1
    )

    if first_proj_claims == 0 or pd.isna(first_proj_claims):
        logger.warning("1st_proj_claims is zero or NaN - defaulting to 1.0")
        return 1.0

    # --- Compute adjust_to_actual_claim ---
    adjust_to_actual_claim = prophet_tranche_vp1 / first_proj_claims

    if pd.isna(adjust_to_actual_claim) or adjust_to_actual_claim == 0:
        logger.warning("adjust_to_actual_claim is zero or NaN - defaulting to 1.0")
        return 1.0

    logger.debug(
        f"adjust_to_actual_claim = {adjust_to_actual_claim:.8f} "
        f"(actual={actual_claims_valdate:.4f}, fv=[{fv_v:.6f},{fv_v1:.6f},{fv_v2:.6f}], "
        f"prophet_tranche=[{prophet_tranche_vp1:.4f},{prophet_tranche_vp2:.4f}])"
    )

    return float(adjust_to_actual_claim)


# =============================================================================
# Main Execution Functions
# =============================================================================

@log_step("Running sensitivity analysis")
def run_sensitivity_analysis(
    client_data: dict,
    prophet_file_path: str,
    valuation_date: str,
    sonia_rates: Dict[int, float],
    num_offsets: int = 501
) -> pd.DataFrame:
    """
    Run sensitivity analysis across all scenarios and offsets.
    
    OPTIMIZED VERSION: Uses pre-computed arrays and vectorized operations
    to achieve 5-10x speedup over the original implementation.
    
    Key optimizations:
    1. Pre-compute all tranche arrays once (avoid repeated pd.to_numeric)
    2. Pre-compute all shifted discount curves per scenario
    3. Use vectorized factor calculations with NumPy arrays
    4. Eliminate DataFrame copies in hot loops
    
    Args:
        client_data: Dict containing tranche data with keys:
            - 'comprehensive_results': Dict of {tranche_name: DataFrame}
        prophet_file_path: Path to Prophet output Excel file
        valuation_date: Valuation date string (YYYY-MM-DD)
        sonia_rates: Dict of {month: sonia_rate}
        num_offsets: Number of offsets to calculate (default 501 for 0-500)
    
    Returns:
        DataFrame with sensitivity results
    """
    from const import (
        RGA_SHARE_BY_DEAL,
        GROSS_UP_FACTOR_BY_DEAL,
        GROSS_UP_FACTOR_DEFERRED,
        GROSS_UP_FACTOR_PENSIONER,
        DEFERRED_PENSIONER_TRANCHES,
        get_deal_name_from_rga_tab,
        DISCOUNT_CURVE_CONFIG,
    )
    
    logger.info(f"Starting OPTIMIZED sensitivity analysis for valuation date: {valuation_date}")
    logger.info(f"Number of scenarios: {len(SCENARIOS)}")
    logger.info(f"Number of offsets: {num_offsets}")
    
    start_time = time.time()
    val_month_end = _get_month_end_timestamp(valuation_date)
    
    results = []
    comprehensive_results = client_data.get('comprehensive_results', {})
    
    logger.info(f"Processing {len(comprehensive_results)} tranches")
    
    # Group tranches by deal
    deal_to_tranches = {}
    for tranche_name in comprehensive_results.keys():
        deal = get_deal_name_from_rga_tab(tranche_name)
        deal_to_tranches.setdefault(deal, []).append(tranche_name)
    
    log_dict_summary(logger, deal_to_tranches, "Deals and Tranches")
    
    # =========================================================================
    # Step 1: Load Prophet data for BE and Stress (1in200) scenarios
    # =========================================================================
    logger.info("Loading Prophet data...")
    prophet_be_df = None
    prophet_stress_df = None
    
    try:
        prophet_be_df = load_prophet_data(prophet_file_path, PROPHET_SHEETS["BE"])
        logger.info(f"Loaded BE Prophet data: {len(prophet_be_df)} rows")
    except Exception as e:
        logger.error(f"Failed to load BE Prophet data: {e}")
        raise
    
    try:
        prophet_stress_df = load_prophet_data(prophet_file_path, PROPHET_SHEETS["1in200"])
        logger.info(f"Loaded 1in200 Prophet data: {len(prophet_stress_df)} rows")
    except Exception as e:
        logger.error(f"Failed to load 1in200 Prophet data: {e}")
        raise
    
    # =========================================================================
    # Step 2: Pre-compute Prophet claims for each deal (RGA share only)
    # =========================================================================
    # 
    # CALCULATION ORDER FOR BLENDED 1IN200 INFLATION PER TRANCHE:
    # -----------------------------------------------------------
    # The calculation follows this logical order (see _precompute_blending_arrays docstring):
    #   1. Blend BE + 1in200 at deal level: blended = (1-bf)*BE + bf*1in200
    #   2. Apply fixed multiplier: multiplier = tranche_claims(prior_month) / total_deal_claims
    #   2.5 Adjust to actual claim: calibrate prophet cf to actual claim at valuation date
    #   3. Inflate per tranche: adjusted = blended / rga_share * (proj_infl / valdate_infl)
    #
    # Implementation applies RGA share here (Step 2), adjust_to_actual_claim in Step 3.5,
    # multiplier and inflation in _precompute_blending_arrays, achieving identical results.
    # =========================================================================
    logger.info("Loading Prophet claims and applying RGA share...")
    
    adjusted_prophet_claims = {}
    
    for deal_name in deal_to_tranches.keys():
        rga_share = RGA_SHARE_BY_DEAL.get(deal_name, 1.0)
        adjusted_prophet_claims[deal_name] = {}
        
        try:
            be_claims_raw = get_prophet_claims_for_deal(prophet_be_df, deal_name)
            be_claims_adj = be_claims_raw / rga_share
            adjusted_prophet_claims[deal_name]['BE'] = be_claims_adj
            logger.debug(f"Deal {deal_name}: Loaded BE claims, {len(be_claims_adj)} dates")
        except Exception as e:
            logger.warning(f"Could not get BE claims for deal {deal_name}: {e}")
            adjusted_prophet_claims[deal_name]['BE'] = pd.Series(dtype=float)
        
        try:
            stress_claims_raw = get_prophet_claims_for_deal(prophet_stress_df, deal_name)
            stress_claims_adj = stress_claims_raw / rga_share
            adjusted_prophet_claims[deal_name]['1in200'] = stress_claims_adj
            logger.debug(f"Deal {deal_name}: Loaded 1in200 claims, {len(stress_claims_adj)} dates")
        except Exception as e:
            logger.warning(f"Could not get 1in200 claims for deal {deal_name}: {e}")
            adjusted_prophet_claims[deal_name]['1in200'] = pd.Series(dtype=float)
    
    # =========================================================================
    # Step 3: Calculate tranche multipliers for each deal
    # =========================================================================
    logger.info("Calculating tranche multipliers...")
    
    deal_multipliers = {}
    for deal_name in deal_to_tranches.keys():
        multipliers = calculate_tranche_multipliers(
            comprehensive_results, valuation_date, deal_name
        )
        deal_multipliers[deal_name] = multipliers
        logger.debug(f"Deal {deal_name}: {len(multipliers)} tranches")
    
    # =========================================================================
    # Step 3.5: Compute adjust_to_actual_claim per tranche (once, reused
    #           across all scenarios/offsets). This calibrates prophet claims
    #           to match the actual observed claim at the valuation date.
    # =========================================================================
    logger.info("Computing adjust_to_actual_claim factors per tranche...")
    
    tranche_adjust_factors: Dict[str, float] = {}
    for deal_name, tranche_list in deal_to_tranches.items():
        be_claims = adjusted_prophet_claims.get(deal_name, {}).get('BE', pd.Series(dtype=float))
        
        for tranche_name in tranche_list:
            tranche_df = comprehensive_results.get(tranche_name)
            multiplier = deal_multipliers.get(deal_name, {}).get(tranche_name, 1.0)
            
            adjust_factor = _compute_adjust_to_actual_claim(
                tranche_df, be_claims, multiplier, valuation_date
            )
            tranche_adjust_factors[tranche_name] = adjust_factor
            
            if adjust_factor != 1.0:
                logger.debug(f"  {tranche_name}: adjust_to_actual_claim = {adjust_factor:.8f}")
    
    logger.info(f"Computed adjust_to_actual_claim for {len(tranche_adjust_factors)} tranches")
    
    # =========================================================================
    # Step 4: OPTIMIZATION - Pre-compute all tranche arrays ONCE
    # This avoids repeated pd.to_numeric() and DataFrame operations
    # =========================================================================
    logger.info("Pre-computing tranche arrays (optimization)...")
    
    tranche_arrays_cache = {}
    tranche_inception_dates = {}
    
    for tranche_name, df in comprehensive_results.items():
        if df is None or df.empty:
            continue
        
        # Pre-compute all arrays for this tranche
        arrays = _prepare_tranche_arrays(df, valuation_date)
        if arrays is not None:
            tranche_arrays_cache[tranche_name] = arrays
            # Get inception date from pre-computed data
            date_dts = arrays['date_datetimes']
            valid_dates = [pd.Timestamp(d) for d in date_dts if pd.notna(d)]
            if valid_dates:
                tranche_inception_dates[tranche_name] = min(valid_dates)
    
    logger.info(f"Pre-computed arrays for {len(tranche_arrays_cache)} tranches")
    
    # =========================================================================
    # Step 4b: Pre-process additional fee vectors (for deals in FEE_TAB_MAPPING)
    # =========================================================================
    fee_vectors_data = client_data.get('fee_vectors_data', {})
    deal_fee_arrays = _prepare_fee_vector_arrays(fee_vectors_data, valuation_date)
    if deal_fee_arrays:
        logger.info(f"Pre-processed additional fee vectors for {len(deal_fee_arrays)} deals: "
                    f"{list(deal_fee_arrays.keys())}")
    
    # =========================================================================
    # Step 5: Run scenarios using OPTIMIZED calculation method
    # =========================================================================
    total_iterations = len(SCENARIOS) * num_offsets * len(deal_to_tranches)
    logger.info(f"Total iterations: {total_iterations:,}")
    
    scenario_tracker = ProgressTracker(
        total=len(SCENARIOS),
        name="Sensitivity Scenarios",
        log_interval=25,
        logger=logger
    )
    
    for scenario in SCENARIOS:
        scenario_start = time.time()
        scenario_name = scenario["name"]
        grading_period = scenario["grading_period"]
        discount_shock = scenario["discount_rate_shock"]
        
        logger.info(f"Processing {scenario_name}: grading_period={grading_period}, "
                   f"discount_shock={discount_shock:.6f}")
        
        # =====================================================================
        # OPTIMIZATION: Pre-compute ALL shifted discount curves for this scenario
        # =====================================================================
        discount_curve_df = calculate_scenario_discount_curve(
            sonia_rates,
            discount_shock,
            max_months=DISCOUNT_CURVE_CONFIG["MAX_MONTHS"] + 1
        )
        shifted_curves = precompute_shifted_curves(discount_curve_df, num_offsets)
        
        # =====================================================================
        # Pre-compute blending arrays for each tranche (base + delta)
        # =====================================================================
        logger.debug(f"  Pre-computing blending arrays for {scenario_name}...")
        
        tranche_blending_data = {}
        for tranche_name, arrays in tranche_arrays_cache.items():
            deal_name = get_deal_name_from_rga_tab(tranche_name)
            multiplier = deal_multipliers.get(deal_name, {}).get(tranche_name, 1.0)
            
            be_claims = adjusted_prophet_claims.get(deal_name, {}).get('BE', pd.Series())
            stress_claims = adjusted_prophet_claims.get(deal_name, {}).get('1in200', pd.Series())
            
            # Get tranche-specific inflation factors
            tranche_infl_lookup, tranche_valdate_infl = _get_tranche_inflation_lookup(
                comprehensive_results, tranche_name, valuation_date
            )
            
            # Pre-compute base and delta arrays (with adjust_to_actual_claim)
            adjust_factor = tranche_adjust_factors.get(tranche_name, 1.0)
            base_arr, delta_arr = _precompute_blending_arrays(
                arrays,
                be_claims,
                stress_claims,
                multiplier,
                tranche_infl_lookup,
                tranche_valdate_infl,
                adjust_to_actual_claim=adjust_factor
            )
            
            tranche_blending_data[tranche_name] = (base_arr, delta_arr)
        
        # =====================================================================
        # Pre-compute blend factors for all offsets (vectorized)
        # =====================================================================
        if grading_period == 0:
            blend_factors = np.zeros(num_offsets)
        else:
            blend_factors = np.minimum(1.0, np.arange(num_offsets) / grading_period)
        
        # =====================================================================
        # Pre-compute offset dates for all offsets
        # =====================================================================
        offset_dates = [get_offset_date(valuation_date, off) for off in range(num_offsets)]
        
        # Pre-compute anchor timestamps for all offsets
        anchor_timestamps = [
            (val_month_end.to_period("M") + off).to_timestamp("M")
            for off in range(num_offsets)
        ]
        
        # =====================================================================
        # Main offset loop - OPTIMIZED
        # =====================================================================
        for offset in range(num_offsets):
            blend_factor = blend_factors[offset]
            shifted_discount = shifted_curves[offset]
            offset_date = offset_dates[offset]
            anchor_date_ts = anchor_timestamps[offset]
            
            for deal_name, tranche_list in deal_to_tranches.items():
                rga_share = RGA_SHARE_BY_DEAL.get(deal_name, 1.0)
                gross_up = GROSS_UP_FACTOR_BY_DEAL.get(deal_name, 0.0)
                gross_up_deferred = GROSS_UP_FACTOR_DEFERRED.get(deal_name, 0.0)
                gross_up_pensioner = GROSS_UP_FACTOR_PENSIONER.get(deal_name, 0.0)
                
                total_pv_premium = 0.0
                total_pv_claims = 0.0
                total_pv_fees = 0.0
                
                for tranche_name in tranche_list:
                    if tranche_name not in tranche_arrays_cache:
                        continue
                    
                    arrays = tranche_arrays_cache[tranche_name]
                    base_arr, delta_arr = tranche_blending_data.get(tranche_name, (None, None))
                    
                    if base_arr is None:
                        continue
                    
                    inception_date = tranche_inception_dates.get(tranche_name)
                    if inception_date is None:
                        continue
                    
                    # Compute blended claims: base + blend_factor * delta
                    blended_claims = base_arr + blend_factor * delta_arr
                    
                    # Get anchor date index for slicing
                    anchor_idx = arrays['date_to_idx'].get(anchor_date_ts, -1)
                    if anchor_idx < 0:
                        # Find closest date >= anchor
                        for i, dt in enumerate(arrays['date_datetimes']):
                            if pd.notna(dt) and pd.Timestamp(dt) >= anchor_date_ts:
                                anchor_idx = i
                                break
                    
                    if anchor_idx < 0 or anchor_idx >= arrays['n_rows']:
                        continue
                    
                    # Compute factors using vectorized method
                    factors = _compute_factors_vectorized(
                        arrays,
                        anchor_date_ts,
                        inception_date,
                        blended_claims
                    )
                    
                    # Compute adjusted float vector using pre-computed arrays
                    adjusted_float_vector = _compute_adjusted_float_vector_fast(
                        arrays,
                        factors,
                        anchor_idx
                    )
                    
                    if len(adjusted_float_vector) == 0:
                        continue
                    
                    # Get sliced arrays from anchor_idx
                    fixed_w_real = arrays['fixed_w_real'][anchor_idx:]
                    deferred_w_real = arrays['deferred_w_real'][anchor_idx:]
                    pensioner_w_real = arrays['pensioner_w_real'][anchor_idx:]
                    
                    # Align lengths
                    curve_len = min(len(adjusted_float_vector), len(shifted_discount), len(fixed_w_real))
                    if curve_len == 0:
                        continue
                    
                    discount_slice = shifted_discount[:curve_len]
                    afv_slice = adjusted_float_vector[:curve_len]
                    fixed_slice = fixed_w_real[:curve_len]
                    deferred_slice = deferred_w_real[:curve_len]
                    pensioner_slice = pensioner_w_real[:curve_len]
                    
                    # Calculate PV metrics (already optimized with NumPy)
                    pv_premium = np.nansum(fixed_slice * discount_slice) * rga_share
                    pv_claims = np.nansum(afv_slice * discount_slice) * rga_share
                    
                    # PV Fees
                    is_dp = tranche_name in DEFERRED_PENSIONER_TRANCHES
                    if is_dp:
                        fee_cf = (deferred_slice * gross_up_deferred + 
                                 pensioner_slice * gross_up_pensioner) * rga_share
                    else:
                        fee_cf = fixed_slice * gross_up * rga_share
                    pv_fees = -np.nansum(fee_cf * discount_slice)
                    
                    total_pv_premium += pv_premium
                    total_pv_claims += pv_claims
                    total_pv_fees += pv_fees
                
                # Add PV of additional fee vectors for this deal (if any)
                if deal_name in deal_fee_arrays:
                    for fee_arr_info in deal_fee_arrays[deal_name]:
                        fee_arr = fee_arr_info['fee_vector']
                        fee_date_to_idx = fee_arr_info['date_to_idx']

                        fee_anchor_idx = fee_date_to_idx.get(anchor_date_ts, -1)
                        if fee_anchor_idx < 0:
                            fee_dts = fee_arr_info['date_datetimes']
                            for fi, fdt in enumerate(fee_dts):
                                if pd.notna(fdt) and pd.Timestamp(fdt) >= anchor_date_ts:
                                    fee_anchor_idx = fi
                                    break

                        if fee_anchor_idx >= 0 and fee_anchor_idx < fee_arr_info['n_rows']:
                            fee_slice = fee_arr[fee_anchor_idx:]
                            min_fee_len = min(len(fee_slice), len(shifted_discount))
                            if min_fee_len > 0:
                                total_pv_fees += -np.nansum(
                                    fee_slice[:min_fee_len] * shifted_discount[:min_fee_len]
                                )
                
                # Calculate exposure and required amount
                exposure = total_pv_claims - total_pv_premium + 0.01 * total_pv_premium
                required_amount = exposure
                
                results.append({
                    "Deal Name": deal_name,
                    "Currency": "GBP",
                    "Scenario": scenario_name,
                    "Date": offset_date,
                    "PV Premium": total_pv_premium,
                    "PV Claims": total_pv_claims,
                    "Exposure": exposure,
                    "PV Fees": total_pv_fees,
                    "Required Amount": required_amount
                })
        
        scenario_elapsed = time.time() - scenario_start
        logger.info(f"Scenario {scenario_name} completed in {scenario_elapsed:.2f}s")
        scenario_tracker.update()
    
    scenario_tracker.complete()
    
    total_elapsed = time.time() - start_time
    logger.info(f"Sensitivity analysis completed in {total_elapsed:.2f}s")
    log_calculation_result(logger, "Total Result Rows", len(results))
    
    return pd.DataFrame(results)


@log_step("Exporting sensitivity results")
def export_sensitivity_results(results_df: pd.DataFrame, output_path: str) -> None:
    """
    Export sensitivity results to Excel file.
    
    Args:
        results_df: DataFrame with sensitivity results
        output_path: Path for output Excel file
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    logger.info(f"Exporting sensitivity results to: {output_path}")
    logger.debug(f"Results shape: {results_df.shape}")
    
    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.debug(f"Created output directory: {output_dir}")
    
    start_time = time.time()
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write each deal to a separate sheet
        unique_deals = results_df["Deal Name"].nunique()
        logger.info(f"Writing {unique_deals} deal sheets")
        
        for deal_name, deal_df in results_df.groupby("Deal Name"):
            sheet_name = deal_name[:31]  # Excel sheet name limit
            deal_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format header
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            ws.freeze_panes = 'A2'
            logger.debug(f"Wrote sheet '{deal_name}': {len(deal_df)} rows")
    
    elapsed = time.time() - start_time
    logger.info(f"Sensitivity results exported in {elapsed:.2f}s: {output_path}")



